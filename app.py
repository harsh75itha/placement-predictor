from flask import (Flask, render_template, request, redirect,
                   url_for, flash, make_response, jsonify)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (LoginManager, UserMixin, login_user,
                         login_required, logout_user, current_user)
from werkzeug.security import generate_password_hash, check_password_hash
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer
from google import genai as new_genai
import pandas as pd
import io
from sklearn.preprocessing import LabelEncoder
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
import os
from datetime import datetime
from functools import wraps

# ============================================================
#  APP & DATABASE CONFIG
# ============================================================
app = Flask(__name__)
# 🔒 Admin secret key — only YOU should know this!
# Change this to whatever secret phrase you want.
ADMIN_SECRET_KEY = os.environ.get('ADMIN_SECRET_KEY','Harshitha_is_the_admin_9900')
app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET', 'placement_predictor_secret_2024')

# ── Flask-Mail config (for Forgot Password emails) ──
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', 'harshithatn75@gmail.com')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')
app.config['MAIL_DEFAULT_SENDER'] = ('Placement Predictor',
                                     os.environ.get('MAIL_USERNAME', 'harshithatn75@gmail.com'))
mail = Mail(app)
# Use DATABASE_URL from environment (Render/cloud).
# If not set, fall back to local MySQL.
db_url = os.environ.get('DATABASE_URL', '').strip()
if db_url:
    # Render gives postgres:// but SQLAlchemy needs postgresql://
    if db_url.startswith('postgres://'):
        db_url = db_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = db_url
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = (
        'mysql+pymysql://root:' + os.environ.get('LOCAL_MYSQL_PASSWORD', 'HaRsH*2005*') + '@localhost/placement_predictor'
    )
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db            = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# ============================================================
#  GEMINI AI
# ============================================================
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', 'AQ.Ab8RN6Lq_mLQQZYXBFJX9bf26E4UgWYbthmER1KLlR-1WuwXLw')
GEMINI_MODEL_NAME = 'gemini-flash-latest'   # free, fast, current
try:
    gemini_client = new_genai.Client(api_key=GEMINI_API_KEY)
    print(f'✅ Gemini AI ready (model: {GEMINI_MODEL_NAME})')
except Exception as e:
    print(f'⚠ Gemini init failed: {e}')
    gemini_client = None

# ============================================================
#  DEGREE DATA
# ============================================================
DEGREE_DATA_SEED = {

    # ── UG GENERAL ────────────────────────────────────────
    'BCA': {
        'full_name': 'Bachelor of Computer Applications',
        'level': 'UG',
        'job_roles': [
            'Software Developer', 'Web Developer', 'Data Analyst',
            'Mobile App Developer', 'System Administrator',
            'Database Administrator', 'UI/UX Designer', 'QA Engineer',
            'Network Engineer', 'IT Support Engineer'
        ],
        'recommended_skills': [
            'Python', 'Java', 'C', 'C++', 'HTML', 'CSS', 'JavaScript',
            'SQL', 'DBMS', 'Data Structures', 'Algorithms', 'OOP',
            'Git', 'Linux', 'React', 'Node.js', 'MongoDB', 'REST APIs'
        ]
    },

    'BBA': {
        'full_name': 'Bachelor of Business Administration',
        'level': 'UG',
        'job_roles': [
            'Business Analyst', 'Marketing Manager', 'HR Manager',
            'Operations Manager', 'Sales Manager', 'Product Manager',
            'Supply Chain Analyst', 'Financial Analyst', 'Entrepreneur',
            'Management Trainee'
        ],
        'recommended_skills': [
            'Excel', 'Financial Statements', 'Marketing Strategy',
            'Digital Marketing', 'Agile', 'Product Strategy',
            'Supply Chain Management', 'Communication', 'Recruitment',
            'Power BI', 'Tableau', 'SQL'
        ]
    },

    'BA': {
        'full_name': 'Bachelor of Arts',
        'level': 'UG',
        'job_roles': [
            'Content Writer', 'Journalist', 'Teacher', 'Civil Services Officer',
            'Public Relations Officer', 'Social Worker', 'HR Executive',
            'Translator', 'Marketing Executive', 'Counselor'
        ],
        'recommended_skills': [
            'Communication', 'Content Writing', 'Research', 'Public Speaking',
            'Digital Marketing', 'MS Office', 'Social Media Management',
            'SEO', 'Creative Writing', 'Data Analysis'
        ]
    },

    'BCom': {
        'full_name': 'Bachelor of Commerce',
        'level': 'UG',
        'job_roles': [
            'Accountant', 'Financial Analyst', 'Auditor', 'Tax Consultant',
            'Banking Officer', 'Investment Analyst', 'Cost Accountant',
            'Finance Manager', 'Insurance Advisor', 'CA Articleship'
        ],
        'recommended_skills': [
            'Financial Statements', 'Tally', 'Excel', 'Financial Modeling',
            'Taxation', 'Auditing', 'SQL', 'Power BI', 'Economics',
            'Communication', 'Accounting'
        ]
    },

    'BEd': {
        'full_name': 'Bachelor of Education',
        'level': 'UG',
        'job_roles': [
            'School Teacher', 'Curriculum Developer', 'Education Counselor',
            'Special Educator', 'Content Developer', 'Tutor',
            'Academic Coordinator', 'E-Learning Designer', 'Principal',
            'Education Administrator'
        ],
        'recommended_skills': [
            'Communication', 'Lesson Planning', 'Classroom Management',
            'Child Psychology', 'MS Office', 'Content Creation',
            'E-Learning Tools', 'Assessment Design', 'Research', 'Public Speaking'
        ]
    },

    'BSc': {
        'full_name': 'Bachelor of Science',
        'level': 'UG',
        'job_roles': [
            'Lab Technician', 'Research Assistant', 'Data Analyst',
            'Quality Control Analyst', 'Science Teacher',
            'Environmental Analyst', 'Pharmaceutical Associate',
            'Biotech Research Associate', 'Clinical Data Manager', 'Medical Coder'
        ],
        'recommended_skills': [
            'Research Methodology', 'Data Analysis', 'MS Excel',
            'Python', 'Statistics', 'Laboratory Techniques',
            'Report Writing', 'Communication', 'SPSS', 'R Programming'
        ]
    },
     # ════════════════════════════════════════════════════════════════
    # NEW BSc BRANCHES + BPharm + BHM (added later)
    # ════════════════════════════════════════════════════════════════

    'BSc CS': {
        'full_name': 'Bachelor of Science in Computer Science',
        'level': 'UG',
        'job_roles': [
            'Software Developer', 'Web Developer', 'Junior Programmer',
            'Application Developer', 'Database Administrator', 'System Analyst',
            'QA Tester', 'Technical Support Engineer', 'Junior Data Analyst',
            'IT Consultant'
        ],
        'recommended_skills': [
            'Python', 'Java', 'C', 'C++', 'HTML', 'CSS', 'JavaScript',
            'SQL', 'Data Structures', 'Algorithms', 'DBMS',
            'Operating Systems', 'Git', 'Linux', 'Computer Networks'
        ]
    },

    'BSc IT': {
        'full_name': 'Bachelor of Science in Information Technology',
        'level': 'UG',
        'job_roles': [
            'IT Support Engineer', 'Web Developer', 'Network Administrator',
            'System Administrator', 'Cybersecurity Analyst',
            'Database Administrator', 'Cloud Support Engineer',
            'ERP Consultant', 'Software Tester', 'Technical Writer'
        ],
        'recommended_skills': [
            'Python', 'Java', 'HTML', 'CSS', 'JavaScript', 'SQL',
            'Networking', 'Cloud Computing', 'Cybersecurity', 'Linux',
            'Windows Server', 'Git', 'DBMS', 'REST APIs'
        ]
    },

    'BSc Mathematics': {
        'full_name': 'Bachelor of Science in Mathematics',
        'level': 'UG',
        'job_roles': [
            'Data Analyst', 'Quantitative Analyst', 'Statistician',
            'Actuarial Analyst', 'Mathematics Teacher', 'Research Assistant',
            'Risk Analyst', 'Operations Research Analyst',
            'Banking Analyst', 'Financial Analyst'
        ],
        'recommended_skills': [
            'Statistics', 'Probability', 'Linear Algebra', 'Calculus',
            'Discrete Mathematics', 'Python', 'R Programming', 'Excel',
            'SQL', 'SPSS', 'Data Analysis', 'Quantitative Aptitude - Arithmetic'
        ]
    },

    'BSc Physics': {
        'full_name': 'Bachelor of Science in Physics',
        'level': 'UG',
        'job_roles': [
            'Research Assistant', 'Physics Teacher', 'Lab Technician',
            'Radiation Safety Officer', 'Astronomical Observer',
            'Quality Control Analyst', 'Junior Scientist (DRDO/ISRO)',
            'Technical Sales Engineer', 'Data Analyst', 'Patent Analyst'
        ],
        'recommended_skills': [
            'Classical Mechanics', 'Quantum Mechanics', 'Electromagnetism',
            'Thermodynamics', 'Optics', 'Mathematical Physics', 'MATLAB',
            'Python', 'Laboratory Techniques', 'Data Analysis',
            'Research Methodology', 'Scientific Communication'
        ]
    },

    'BSc Chemistry': {
        'full_name': 'Bachelor of Science in Chemistry',
        'level': 'UG',
        'job_roles': [
            'Chemical Lab Analyst', 'Quality Control Chemist',
            'Pharmaceutical Sales Representative', 'Research Associate',
            'Chemistry Teacher', 'Environmental Chemist',
            'Food Quality Analyst', 'Production Chemist',
            'Forensic Chemist', 'Patent Officer'
        ],
        'recommended_skills': [
            'Organic Chemistry', 'Inorganic Chemistry', 'Physical Chemistry',
            'Analytical Chemistry', 'Biochemistry', 'Laboratory Techniques',
            'Chromatography (HPLC / GC)', 'Spectroscopy', 'Quality Control',
            'MS Excel', 'Research Methodology', 'Scientific Writing'
        ]
    },

    'BSc Biology': {
        'full_name': 'Bachelor of Science in Biological Sciences',
        'level': 'UG',
        'job_roles': [
            'Biological Lab Technician', 'Research Assistant',
            'Biology Teacher', 'Healthcare Associate',
            'Pharmaceutical Sales Representative',
            'Quality Control Analyst (Pharma)',
            'Clinical Research Coordinator', 'Junior Biotechnologist',
            'Environmental Consultant', 'Medical Coder'
        ],
        'recommended_skills': [
            'Microbiology', 'Cell Biology', 'Genetics', 'Biochemistry',
            'Anatomy & Physiology', 'Ecology', 'Molecular Biology',
            'Laboratory Techniques', 'Data Analysis', 'MS Excel',
            'Research Methodology', 'Scientific Writing'
        ]
    },

    'BSc Microbiology': {
        'full_name': 'Bachelor of Science in Microbiology',
        'level': 'UG',
        'job_roles': [
            'Microbiologist', 'Lab Technician',
            'Quality Control Microbiologist', 'Food Safety Analyst',
            'Pharmaceutical Microbiologist', 'Research Associate',
            'Clinical Laboratory Scientist', 'Biomedical Researcher',
            'Microbiology Teacher', 'Environmental Microbiologist'
        ],
        'recommended_skills': [
            'Microbial Cultivation Techniques', 'Sterilization Methods',
            'Microscopy Techniques', 'Microbial Identification',
            'PCR Techniques', 'Microbial Genetics', 'Immunology',
            'Bioinformatics', 'Laboratory Techniques', 'Quality Control',
            'Research Methodology', 'MS Excel', 'Scientific Writing'
        ]
    },

    'BSc Biotechnology': {
        'full_name': 'Bachelor of Science in Biotechnology',
        'level': 'UG',
        'job_roles': [
            'Biotech Research Associate', 'Quality Control Officer',
            'Clinical Research Associate', 'Bioprocess Engineer',
            'Bioinformatics Analyst', 'R&D Associate (Pharma)',
            'Biotech Sales Executive', 'Lab Technician',
            'Tissue Culture Specialist', 'Genomics Analyst'
        ],
        'recommended_skills': [
            'Molecular Biology', 'Genetic Engineering',
            'Cell Culture Techniques', 'PCR & Gel Electrophoresis',
            'Bioinformatics', 'Microbiology', 'Biochemistry',
            'Immunology', 'Fermentation Technology', 'Python',
            'Laboratory Techniques', 'Research Methodology',
            'Scientific Writing'
        ]
    },

    'BSc Nursing': {
        'full_name': 'Bachelor of Science in Nursing',
        'level': 'UG',
        'job_roles': [
            'Staff Nurse', 'Critical Care Nurse',
            'Operation Theatre Nurse', 'Community Health Nurse',
            'Pediatric Nurse', 'Nursing Tutor', 'Home Care Nurse',
            'Clinical Nurse Manager', 'Public Health Nurse', 'Nurse Educator'
        ],
        'recommended_skills': [
            'Patient Care', 'Anatomy & Physiology', 'Pharmacology',
            'Medical-Surgical Nursing', 'First Aid & Emergency Care',
            'Infection Control', 'Vital Signs Monitoring',
            'Medical Documentation', 'Communication',
            'Empathy & Counseling', 'Nutrition', 'Community Health',
            'Hospital Information Systems'
        ]
    },

    'BSc Agriculture': {
        'full_name': 'Bachelor of Science in Agriculture',
        'level': 'UG',
        'job_roles': [
            'Agricultural Officer', 'Farm Manager', 'Agronomist',
            'Agricultural Sales Representative', 'Plant Breeder',
            'Soil Scientist', 'Horticulturist',
            'Agricultural Banking Officer (NABARD)',
            'Agricultural Extension Officer', 'Agribusiness Manager'
        ],
        'recommended_skills': [
            'Soil Science', 'Crop Production', 'Plant Pathology',
            'Horticulture', 'Agronomy', 'Irrigation Management',
            'Agricultural Economics', 'Pest Management', 'Agribusiness',
            'GIS for Agriculture', 'MS Excel', 'Communication',
            'Field Survey Techniques'
        ]
    },

    'BSc Data Science': {
        'full_name': 'Bachelor of Science in Data Science',
        'level': 'UG',
        'job_roles': [
            'Data Analyst', 'Junior Data Scientist',
            'Business Intelligence Analyst', 'Data Engineer',
            'Machine Learning Engineer',
            'Data Visualization Specialist', 'Quantitative Analyst',
            'Research Analyst', 'Statistician', 'Marketing Analyst'
        ],
        'recommended_skills': [
            'Python', 'R Programming', 'SQL', 'Statistics',
            'Machine Learning', 'Data Visualization', 'Pandas',
            'NumPy', 'Power BI', 'Tableau', 'Excel', 'Data Cleaning',
            'Big Data Basics', 'Linear Algebra', 'Communication'
        ]
    },

    'BSc Psychology': {
        'full_name': 'Bachelor of Science in Psychology',
        'level': 'UG',
        'job_roles': [
            'Counsellor', 'School Psychologist', 'HR Executive',
            'Mental Health Worker', 'Behavioural Therapist',
            'Research Assistant', 'Rehabilitation Counsellor',
            'Social Worker', 'Career Counsellor', 'UX Researcher'
        ],
        'recommended_skills': [
            'Counselling Techniques', 'Cognitive Psychology',
            'Developmental Psychology', 'Abnormal Psychology',
            'Social Psychology', 'Psychological Assessment',
            'Statistics', 'SPSS', 'Research Methodology',
            'Communication', 'Empathy & Active Listening',
            'Case Study Writing', 'Report Writing'
        ]
    },

    'BSc Forensic': {
        'full_name': 'Bachelor of Science in Forensic Science',
        'level': 'UG',
        'job_roles': [
            'Forensic Scientist', 'Crime Scene Investigator',
            'Forensic Laboratory Analyst', 'Forensic DNA Analyst',
            'Forensic Toxicologist', 'Fingerprint Expert',
            'Forensic Ballistics Expert', 'Cyber Forensic Analyst',
            'Forensic Document Examiner', 'Forensic Pathologist Assistant'
        ],
        'recommended_skills': [
            'Crime Scene Investigation', 'Forensic Photography',
            'DNA Analysis', 'Fingerprint Analysis', 'Toxicology',
            'Ballistics & Firearm Analysis', 'Forensic Chemistry',
            'Forensic Biology', 'Indian Evidence Act',
            'Criminal Law (IPC, CrPC, BNS)', 'Microscopy Techniques',
            'Document Examination', 'Cyber Forensics', 'Report Writing',
            'Chain of Custody Procedures'
        ]
    },

    'BPharm': {
        'full_name': 'Bachelor of Pharmacy',
        'level': 'UG',
        'job_roles': [
            'Pharmacist', 'Hospital Pharmacist', 'Community Pharmacist',
            'Clinical Research Associate',
            'Pharmaceutical Sales Representative', 'Drug Inspector',
            'Pharmaceutical Production Officer',
            'Quality Control Officer (Pharma)', 'Medical Writer',
            'Drug Safety Associate'
        ],
        'recommended_skills': [
            'Pharmacology', 'Pharmaceutics', 'Pharmacognosy',
            'Medicinal Chemistry', 'Pharmaceutical Analysis',
            'Clinical Pharmacy', 'Hospital Pharmacy',
            'Patient Counseling', 'Drug Safety', 'Quality Assurance',
            'Pharmaceutical Marketing', 'Drug Regulatory Affairs',
            'Communication'
        ]
    },

    'BHM': {
        'full_name': 'Bachelor of Hotel Management',
        'level': 'UG',
        'job_roles': [
            'Hotel Manager', 'Front Office Executive', 'F&B Manager',
            'Chef', 'Housekeeping Manager', 'Event Coordinator',
            'Restaurant Manager', 'Travel & Tourism Manager',
            'Hospitality Trainer', 'Cruise Liner Staff'
        ],
        'recommended_skills': [
            'Food Production', 'F&B Service', 'Front Office Management',
            'Housekeeping Management', 'Hotel Accounts',
            'Tourism Management', 'Event Management', 'Customer Service',
            'Culinary Arts', 'Beverage Knowledge', 'Hospitality Marketing',
            'Communication', 'MS Office', 'Hospitality Law'
        ]
    },


    # ── UG ENGINEERING ────────────────────────────────────
    'BE/BTech - CSE': {
        'full_name': 'B.E / B.Tech - Computer Science Engineering',
        'level': 'UG Engineering',
        'job_roles': [
            'Software Engineer', 'Full Stack Developer', 'Data Scientist',
            'DevOps Engineer', 'Cloud Engineer', 'Cybersecurity Analyst',
            'AI/ML Engineer', 'System Analyst', 'Database Administrator',
            'Product Manager'
        ],
        'recommended_skills': [
            'Python', 'Java', 'C++', 'Data Structures', 'Algorithms',
            'DBMS', 'Operating Systems', 'Computer Networks', 'OOP',
            'React', 'Node.js', 'SQL', 'Git', 'Docker', 'AWS',
            'Machine Learning', 'Linux', 'REST APIs'
        ]
    },

    'BE/BTech - ECE': {
        'full_name': 'B.E / B.Tech - Electronics & Communication Engineering',
        'level': 'UG Engineering',
        'job_roles': [
            'Embedded Systems Engineer', 'VLSI Design Engineer', 'RF Engineer',
            'IoT Developer', 'Telecom Engineer', 'Signal Processing Engineer',
            'Hardware Design Engineer', 'PCB Design Engineer',
            'Network Engineer', 'R&D Engineer'
        ],
        'recommended_skills': [
            'Embedded C', 'VLSI', 'Verilog', 'MATLAB', 'PCB Design',
            'Signal Processing', 'Microcontrollers', 'Arduino',
            'Raspberry Pi', 'Communication Systems', 'Circuit Design',
            'Python', 'LabVIEW', 'Antenna Design', 'IoT'
        ]
    },

    'BE/BTech - EEE': {
        'full_name': 'B.E / B.Tech - Electrical & Electronics Engineering',
        'level': 'UG Engineering',
        'job_roles': [
            'Electrical Engineer', 'Power Systems Engineer',
            'Control Systems Engineer', 'PLC Programmer', 'SCADA Engineer',
            'Solar Energy Engineer', 'Automation Engineer',
            'Maintenance Engineer', 'Instrumentation Engineer',
            'Electrical Design Engineer'
        ],
        'recommended_skills': [
            'Power Systems', 'Control Systems', 'PLC Programming',
            'SCADA', 'AutoCAD Electrical', 'MATLAB', 'Circuit Design',
            'Electrical Machines', 'Power Electronics', 'Switchgear',
            'Solar PV Systems', 'Protection Systems'
        ]
    },

    'BE/BTech - Mechanical': {
        'full_name': 'B.E / B.Tech - Mechanical Engineering',
        'level': 'UG Engineering',
        'job_roles': [
            'Mechanical Design Engineer', 'Production Engineer',
            'Quality Engineer', 'Maintenance Engineer', 'HVAC Engineer',
            'Automobile Engineer', 'Manufacturing Engineer',
            'R&D Engineer', 'Project Engineer', 'CAD/CAM Engineer'
        ],
        'recommended_skills': [
            'AutoCAD', 'SolidWorks', 'CATIA', 'ANSYS', 'CFD',
            'Thermodynamics', 'Fluid Mechanics', 'Heat Transfer',
            'Manufacturing Processes', 'GD&T', 'MATLAB',
            'Finite Element Analysis', 'Lean Manufacturing',
            'Six Sigma', 'Project Management'
        ]
    },

    'BE/BTech - Civil': {
        'full_name': 'B.E / B.Tech - Civil Engineering',
        'level': 'UG Engineering',
        'job_roles': [
            'Site Engineer', 'Structural Engineer', 'Estimation Engineer',
            'Project Manager', 'Surveyor', 'Quality Control Engineer',
            'Environmental Engineer', 'Transportation Engineer',
            'Geotechnical Engineer', 'Urban Planner'
        ],
        'recommended_skills': [
            'AutoCAD', 'STAAD Pro', 'ETABS', 'Revit', 'MS Project',
            'Structural Analysis', 'Surveying', 'Estimation & Costing',
            'Concrete Technology', 'Soil Mechanics', 'GIS',
            'Construction Management', 'Quantity Surveying'
        ]
    },

    'BE/BTech - IT': {
        'full_name': 'B.E / B.Tech - Information Technology',
        'level': 'UG Engineering',
        'job_roles': [
            'Software Developer', 'Web Developer', 'Network Administrator',
            'Database Administrator', 'Cybersecurity Analyst',
            'Cloud Engineer', 'IT Support Engineer', 'Systems Analyst',
            'Business Intelligence Analyst', 'ERP Consultant'
        ],
        'recommended_skills': [
            'Python', 'Java', 'HTML', 'CSS', 'JavaScript', 'SQL',
            'Networking', 'Cybersecurity', 'Cloud Computing',
            'DBMS', 'Linux', 'Git', 'React', 'PHP',
            'Data Analysis', 'REST APIs'
        ]
    },

    'BE/BTech - Chemical': {
        'full_name': 'B.E / B.Tech - Chemical Engineering',
        'level': 'UG Engineering',
        'job_roles': [
            'Process Engineer', 'Chemical Plant Engineer', 'Quality Analyst',
            'Safety Engineer', 'Environmental Engineer', 'R&D Engineer',
            'Production Engineer', 'Petroleum Engineer',
            'Food Technology Engineer', 'Pharmaceutical Engineer'
        ],
        'recommended_skills': [
            'Process Simulation', 'HYSYS', 'MATLAB', 'AutoCAD',
            'Chemical Process Design', 'Mass Transfer', 'Heat Transfer',
            'Thermodynamics', 'Safety Management', 'Quality Control',
            'Six Sigma', 'Process Optimization'
        ]
    },

    'BE/BTech - Aerospace': {
        'full_name': 'B.E / B.Tech - Aerospace Engineering',
        'level': 'UG Engineering',
        'job_roles': [
            'Aerospace Design Engineer', 'Avionics Engineer',
            'Propulsion Engineer', 'Structural Analyst',
            'Flight Test Engineer', 'Navigation Systems Engineer',
            'Spacecraft Engineer', 'Defence R&D Engineer',
            'UAV Engineer', 'Simulation Engineer'
        ],
        'recommended_skills': [
            'MATLAB', 'ANSYS', 'CFD', 'SolidWorks', 'AutoCAD',
            'Aerodynamics', 'Propulsion Systems', 'Avionics',
            'Flight Mechanics', 'Structural Analysis',
            'Python', 'Control Systems', 'Finite Element Analysis'
        ]
    },

    # ── PG ────────────────────────────────────────────────
    'MCA': {
        'full_name': 'Master of Computer Applications',
        'level': 'PG',
        'job_roles': [
            'Senior Software Engineer', 'Full Stack Developer',
            'Data Scientist', 'Cloud Engineer', 'DevOps Engineer',
            'Machine Learning Engineer', 'Cybersecurity Analyst',
            'Solution Architect', 'Project Manager', 'AI Engineer'
        ],
        'recommended_skills': [
            'Python', 'Java', 'Machine Learning', 'Deep Learning',
            'AWS', 'Docker', 'Kubernetes', 'React', 'Node.js', 'SQL',
            'MongoDB', 'Git', 'Linux', 'REST APIs', 'Data Structures',
            'Algorithms', 'NLP', 'TensorFlow'
        ]
    },

    'MBA': {
        'full_name': 'Master of Business Administration',
        'level': 'PG',
        'job_roles': [
            'Business Analyst', 'Product Manager', 'Marketing Manager',
            'Finance Manager', 'HR Manager', 'Operations Manager',
            'Strategy Consultant', 'Investment Banker', 'Brand Manager',
            'Management Consultant'
        ],
        'recommended_skills': [
            'Financial Modeling', 'Business Strategy', 'Marketing Analytics',
            'Excel', 'Power BI', 'SQL', 'Agile', 'Leadership',
            'Communication', 'Data Analysis', 'Tableau', 'CRM Tools',
            'Supply Chain Management', 'Digital Marketing'
        ]
    },

    'MTech': {
        'full_name': 'Master of Technology',
        'level': 'PG',
        'job_roles': [
            'Research Engineer', 'Senior Software Engineer', 'Data Scientist',
            'Embedded Systems Engineer', 'VLSI Design Engineer', 'AI Researcher',
            'Cloud Architect', 'Robotics Engineer', 'Cybersecurity Engineer',
            'R&D Engineer'
        ],
        'recommended_skills': [
            'Python', 'Machine Learning', 'Deep Learning', 'MATLAB',
            'C++', 'Verilog', 'VLSI', 'Signal Processing', 'TensorFlow',
            'PyTorch', 'AWS', 'Docker', 'Linux', 'Research',
            'Data Structures', 'Algorithms'
        ]
    },

    'MEd': {
        'full_name': 'Master of Education',
        'level': 'PG',
        'job_roles': [
            'College Lecturer', 'Education Researcher', 'Curriculum Designer',
            'School Principal', 'Education Policy Analyst',
            'Instructional Designer', 'Training & Development Manager',
            'Academic Dean', 'Special Education Director', 'E-Learning Specialist'
        ],
        'recommended_skills': [
            'Research Methodology', 'Curriculum Design', 'Educational Technology',
            'Assessment & Evaluation', 'Leadership', 'Communication',
            'Data Analysis', 'E-Learning Tools', 'Psychology', 'Content Development'
        ]
    },

    'MA': {
        'full_name': 'Master of Arts',
        'level': 'PG',
        'job_roles': [
            'College Lecturer', 'Researcher', 'Journalist', 'Content Strategist',
            'Civil Services Officer', 'Policy Analyst', 'HR Manager',
            'Social Worker', 'Translator', 'Public Relations Manager'
        ],
        'recommended_skills': [
            'Research Methodology', 'Academic Writing', 'Content Writing',
            'Public Speaking', 'Data Analysis', 'Communication', 'MS Office',
            'SEO', 'Social Media Management', 'Critical Thinking'
        ]
    },

    'MCom': {
        'full_name': 'Master of Commerce',
        'level': 'PG',
        'job_roles': [
            'Senior Financial Analyst', 'Investment Banker', 'Chief Accountant',
            'Finance Controller', 'Risk Manager', 'Portfolio Manager',
            'Tax Advisor', 'Forensic Accountant', 'Commerce Lecturer',
            'CFO (Career Path)'
        ],
        'recommended_skills': [
            'Financial Modeling', 'Financial Statements', 'Excel', 'Tally',
            'Bloomberg Terminal', 'Risk Analysis', 'Taxation', 'Auditing',
            'Power BI', 'SQL', 'Investment Analysis', 'Corporate Finance'
        ]
    },

    'MSc': {
        'full_name': 'Master of Science',
        'level': 'PG',
        'job_roles': [
            'Research Scientist', 'Data Scientist', 'Lab Manager',
            'Quality Assurance Manager', 'Science Lecturer',
            'Environmental Consultant', 'Pharmaceutical Scientist',
            'Biotech Researcher', 'Clinical Research Associate',
            'Forensic Scientist'
        ],
        'recommended_skills': [
            'Research Methodology', 'Python', 'R Programming', 'Statistics',
            'SPSS', 'Data Analysis', 'Machine Learning', 'Laboratory Techniques',
            'Academic Writing', 'MS Excel', 'Scientific Communication'
        ]
    },

    # ── COMPETITIVE EXAMS ─────────────────────────────────
    'UPSC - Civil Services': {
        'full_name': 'UPSC Civil Services Examination',
        'level': 'Competitive Exam',
        'job_roles': [
            'IAS Officer', 'IPS Officer', 'IFS Officer', 'IRS Officer',
            'IFoS Officer', 'IDAS Officer', 'IRTS Officer',
            'Central Secretariat Service', 'Defence Accounts Service',
            'Indian Postal Service'
        ],
        'recommended_skills': [
            'General Studies - History', 'General Studies - Geography',
            'General Studies - Polity & Constitution',
            'General Studies - Indian Economy',
            'General Studies - Environment & Ecology',
            'General Studies - Science & Technology',
            'General Studies - International Relations',
            'Current Affairs (The Hindu / Indian Express)',
            'CSAT - Aptitude & Reasoning',
            'Essay Writing', 'Ethics & Integrity',
            'Answer Writing Practice', 'Optional Subject Mastery',
            'Mind Mapping & Notes Making',
            'Previous Year UPSC Papers'
        ]
    },

    'KPSC - Karnataka PSC': {
        'full_name': 'Karnataka Public Service Commission Examination',
        'level': 'Competitive Exam',
        'job_roles': [
            'Karnataka Administrative Service (KAS)',
            'Karnataka Police Service (KPS)',
            'Deputy Superintendent of Police',
            'Assistant Commissioner',
            'Commercial Tax Officer',
            'Treasury Officer',
            'Assistant Director (Various Depts)',
            'Revenue Inspector',
            'Block Development Officer',
            'Karnataka Forest Service'
        ],
        'recommended_skills': [
            'Karnataka History & Culture',
            'Karnataka Geography',
            'Karnataka Polity & Administration',
            'General Studies (National Level)',
            'Current Affairs - Karnataka & National',
            'Kannada Language & Grammar',
            'Indian Constitution',
            'Economy of Karnataka',
            'Science & Technology',
            'Mental Ability & Reasoning',
            'Essay & Answer Writing',
            'Previous Year KPSC Papers',
            'Karnataka Acts & Rules'
        ]
    },

    'Railway Exams - RRB': {
        'full_name': 'Railway Recruitment Board Examinations (RRB NTPC / JE / Group D)',
        'level': 'Competitive Exam',
        'job_roles': [
            'Junior Engineer (JE)', 'Senior Section Engineer (SSE)',
            'Station Master', 'Goods Guard', 'Loco Pilot',
            'Technician Grade III', 'RPF Constable / SI',
            'NTPC - Clerk / Typist / Commercial Apprentice',
            'Group D - Track Maintainer / Helper',
            'ALP - Assistant Loco Pilot'
        ],
        'recommended_skills': [
            'Mathematics - Arithmetic & Algebra',
            'General Intelligence & Reasoning',
            'General Science - Physics',
            'General Science - Chemistry',
            'General Science - Biology',
            'General Awareness & Current Affairs',
            'Technical Subjects (ECE / EEE / Mech / Civil as per post)',
            'Basic Computer Knowledge',
            'English / Hindi Language',
            'Speed & Accuracy Practice',
            'Previous Year RRB Papers',
            'Typing Speed (for NTPC posts)'
        ]
    },

    'Bank Exams - IBPS/SBI': {
        'full_name': 'Banking Examinations (IBPS PO/Clerk, SBI PO/Clerk, RBI Grade B)',
        'level': 'Competitive Exam',
        'job_roles': [
            'Probationary Officer (PO)', 'Bank Clerk',
            'Specialist Officer - IT / HR / Law / Marketing',
            'RBI Grade B Officer', 'RBI Assistant',
            'NABARD Development Assistant',
            'SBI Junior Associate', 'SBI PO',
            'IBPS RRB Officer',
            'Insurance Officer (LIC / GIC)'
        ],
        'recommended_skills': [
            'Quantitative Aptitude - Number System & DI',
            'Quantitative Aptitude - Simplification & Percentages',
            'Reasoning Ability - Puzzles & Seating Arrangement',
            'Reasoning Ability - Syllogism & Coding-Decoding',
            'English Language - Reading Comprehension',
            'English Language - Grammar & Vocabulary',
            'General Awareness - Banking & Economy',
            'General Awareness - Current Affairs',
            'Computer Knowledge - MS Office & Internet',
            'Financial Awareness - RBI Policies & Budget',
            'Data Interpretation',
            'Speed Maths',
            'Mock Test Practice',
            'Descriptive Writing (for PO Mains)',
            'Previous Year Bank Papers'
        ]
    },

    'SSC Exams': {
        'full_name': 'Staff Selection Commission Examinations (SSC CGL / CHSL / MTS)',
        'level': 'Competitive Exam',
        'job_roles': [
            'Income Tax Inspector',
            'Excise Inspector',
            'Assistant Section Officer (MEA / CBI / NIA)',
            'Sub-Inspector CBI',
            'Auditor / Accountant (CAG / CGDA)',
            'Statistical Investigator',
            'Tax Assistant',
            'Lower Division Clerk (LDC)',
            'Data Entry Operator',
            'Multi-Tasking Staff (MTS)'
        ],
        'recommended_skills': [
            'Quantitative Aptitude - Arithmetic',
            'Quantitative Aptitude - Mensuration & Geometry',
            'Quantitative Aptitude - Algebra & Trigonometry',
            'General Intelligence & Reasoning',
            'English Language & Comprehension',
            'General Awareness - History & Polity',
            'General Awareness - Geography & Economy',
            'Current Affairs',
            'Computer Proficiency',
            'Typing Speed (for LDC / DEO)',
            'Data Interpretation',
            'Mock Tests & Speed Practice',
            'Previous Year SSC Papers'
        ]
    },

}

# These globals get populated from DB at startup and refreshed after admin changes
DEGREE_DATA = {}
ALL_DEGREES = []
COMPETITIVE_EXAM_ROLES = []

def get_skills_for_role(role, default_skills):
    """Find the right skill set for a given role.
    If the role belongs to a competitive exam, return that exam's skills.
    Otherwise return the student's degree skills.
    """
    for _name, _data in DEGREE_DATA.items():
        if role in _data.get('job_roles', []):
            return _data.get('recommended_skills', default_skills)
    return default_skills

# ============================================================
#  RECOMMENDATIONS
# ============================================================
RECOMMENDATIONS = {
    "Python":             {"course": "Python for Everybody - Coursera",                "platform": "https://www.coursera.org",              "project": "Build a calculator or to-do app",              "duration": "4 weeks"},
    "SQL":                {"course": "MySQL for Beginners - Udemy",                    "platform": "https://www.udemy.com",                 "project": "Build a Student Database System",              "duration": "2-3 weeks"},
    "Machine Learning":   {"course": "ML Specialization - Coursera (Andrew Ng)",       "platform": "https://www.coursera.org",              "project": "Build a house price prediction model",         "duration": "8 weeks"},
    "Data Structures":    {"course": "DSA in Python - Udemy",                          "platform": "https://www.udemy.com",                 "project": "Solve 30 LeetCode problems",                   "duration": "6-8 weeks"},
    "Java":               {"course": "Java Programming Masterclass - Udemy",           "platform": "https://www.udemy.com",                 "project": "Build a simple banking system",                "duration": "6 weeks"},
    "HTML":               {"course": "HTML & CSS Full Course - freeCodeCamp",          "platform": "https://www.youtube.com",               "project": "Build your personal portfolio website",        "duration": "2 weeks"},
    "CSS":                {"course": "CSS Full Course - freeCodeCamp",                 "platform": "https://www.youtube.com",               "project": "Build a responsive landing page",              "duration": "2 weeks"},
    "JavaScript":         {"course": "JavaScript Full Course - freeCodeCamp",          "platform": "https://www.youtube.com",               "project": "Build a weather app",                          "duration": "4 weeks"},
    "React":              {"course": "React JS - Traversy Media YouTube",              "platform": "https://www.youtube.com",               "project": "Build a todo app using React",                 "duration": "4 weeks"},
    "Node.js":            {"course": "Node.js Full Course - Traversy Media",           "platform": "https://www.youtube.com",               "project": "Build a REST API with Express",                "duration": "3 weeks"},
    "MongoDB":            {"course": "MongoDB for Beginners - Udemy",                  "platform": "https://www.udemy.com",                 "project": "Build a blog with MongoDB",                    "duration": "2 weeks"},
    "Git":                {"course": "Git & GitHub - freeCodeCamp YouTube",            "platform": "https://www.youtube.com",               "project": "Push your project to GitHub",                  "duration": "1 week"},
    "Docker":             {"course": "Docker for Beginners - Udemy",                   "platform": "https://www.udemy.com",                 "project": "Dockerize a Flask app",                        "duration": "2-3 weeks"},
    "AWS":                {"course": "AWS Cloud Practitioner - Udemy",                 "platform": "https://www.udemy.com",                 "project": "Deploy a web app on AWS EC2",                  "duration": "4-6 weeks"},
    "Linux":              {"course": "Linux Command Line Basics - Udemy",              "platform": "https://www.udemy.com",                 "project": "Set up a Linux server",                        "duration": "2 weeks"},
    "Deep Learning":      {"course": "Deep Learning Specialization - Coursera",        "platform": "https://www.coursera.org",              "project": "Build an image classifier using CNN",          "duration": "6 weeks"},
    "TensorFlow":         {"course": "TensorFlow Developer Certificate - Coursera",    "platform": "https://www.coursera.org",              "project": "Build a digit recognizer",                     "duration": "4 weeks"},
    "Power BI":           {"course": "Power BI Full Course - Simplilearn YouTube",     "platform": "https://www.youtube.com",               "project": "Build an HR Analytics Dashboard",              "duration": "2 weeks"},
    "Excel":              {"course": "Microsoft Excel for Beginners - Udemy",          "platform": "https://www.udemy.com",                 "project": "Create a monthly budget tracker",              "duration": "1-2 weeks"},
    "Tableau":            {"course": "Tableau for Beginners - Udemy",                  "platform": "https://www.udemy.com",                 "project": "Build a Covid-19 dashboard",                   "duration": "2 weeks"},
    "Communication":      {"course": "Business Communication - Coursera",              "platform": "https://www.coursera.org",              "project": "Give a 5-min presentation on your project",    "duration": "2 weeks"},
    "Digital Marketing":  {"course": "Google Digital Marketing Certificate - Coursera","platform": "https://www.coursera.org",              "project": "Run a Google Ads campaign",                    "duration": "4 weeks"},
    "Financial Modeling": {"course": "Financial Modeling - Udemy",                     "platform": "https://www.udemy.com",                 "project": "Build a 3-statement financial model",          "duration": "3 weeks"},
    "Tally":              {"course": "Tally Prime Full Course - YouTube",              "platform": "https://www.youtube.com",               "project": "Create a full accounting ledger",              "duration": "2 weeks"},
    "Taxation":           {"course": "Income Tax & GST - Udemy",                       "platform": "https://www.udemy.com",                 "project": "File a mock GST return",                       "duration": "2 weeks"},
    "SEO":                {"course": "SEO Training - Moz Beginner's Guide",            "platform": "https://moz.com/beginners-guide-to-seo","project": "Optimize a blog for SEO",                     "duration": "2 weeks"},
    "Content Writing":    {"course": "Content Marketing - HubSpot Academy",            "platform": "https://academy.hubspot.com",           "project": "Write 5 SEO blog articles",                    "duration": "2 weeks"},
    "Research Methodology": {"course": "Research Methods - Coursera",                  "platform": "https://www.coursera.org",              "project": "Write a 15-page research paper",               "duration": "3 weeks"},
    "Curriculum Design":  {"course": "Instructional Design - Coursera",                "platform": "https://www.coursera.org",              "project": "Design a full semester curriculum",            "duration": "3 weeks"},
    "DBMS":               {"course": "Database Management Systems - NPTEL",            "platform": "https://nptel.ac.in",                   "project": "Design a college management database",         "duration": "3 weeks"},
    "OOP":                {"course": "OOP in Python - Udemy",                          "platform": "https://www.udemy.com",                 "project": "Build a library management system",            "duration": "2 weeks"},
    "Algorithms":         {"course": "Algorithms Specialization - Coursera (Stanford)","platform": "https://www.coursera.org",              "project": "Solve 50 LeetCode problems",                   "duration": "6 weeks"},
    "REST APIs":          {"course": "REST API Design - Udemy",                        "platform": "https://www.udemy.com",                 "project": "Build a student management API",               "duration": "2 weeks"},
    "Kubernetes":         {"course": "Kubernetes for Beginners - Udemy",               "platform": "https://www.udemy.com",                 "project": "Deploy a microservices app",                   "duration": "3 weeks"},
    "NLP":                {"course": "Natural Language Processing - Coursera",         "platform": "https://www.coursera.org",              "project": "Build a sentiment analysis model",             "duration": "4 weeks"},
    "C":                  {"course": "C Programming for Beginners - Udemy",            "platform": "https://www.udemy.com",                 "project": "Build a student record system in C",           "duration": "3 weeks"},
    "C++":                {"course": "C++ Programming - Udemy",                        "platform": "https://www.udemy.com",                 "project": "Build a library management system in C++",     "duration": "4 weeks"},
    "Agile":              {"course": "Agile Project Management - Coursera",            "platform": "https://www.coursera.org",              "project": "Manage a project using Agile",                 "duration": "2 weeks"},
    "MS Office":          {"course": "Microsoft Office Complete - Udemy",              "platform": "https://www.udemy.com",                 "project": "Create a project report using Word & PPT",     "duration": "1 week"},
    "Financial Statements": {"course": "Financial Accounting - Coursera",             "platform": "https://www.coursera.org",              "project": "Analyze a real company's financials",          "duration": "3 weeks"},
    "Investment Analysis": {"course": "Investment Analysis - Coursera (Yale)",         "platform": "https://www.coursera.org",              "project": "Compare 3 mutual funds",                       "duration": "3 weeks"},
    "MATLAB":             {"course": "MATLAB Onramp - MathWorks Free",                 "platform": "https://matlabacademy.mathworks.com",   "project": "Solve engineering problems in MATLAB",         "duration": "2 weeks"},
    "PyTorch":            {"course": "PyTorch for Deep Learning - Udemy",              "platform": "https://www.udemy.com",                 "project": "Build an image classifier",                    "duration": "4 weeks"},
    "Data Visualization": {"course": "Data Visualization with Python - Coursera",      "platform": "https://www.coursera.org",              "project": "Create a sales dashboard",                     "duration": "2 weeks"},
    "Pandas":             {"course": "Pandas for Data Analysis - Kaggle",              "platform": "https://www.kaggle.com/learn/pandas",  "project": "Clean and analyze a messy dataset",            "duration": "1-2 weeks"},
    "AutoCAD":            {"course": "AutoCAD 2024 Complete Course - Udemy",           "platform": "https://www.udemy.com",                 "project": "Design a mechanical part drawing",             "duration": "3 weeks"},
    "SolidWorks":         {"course": "SolidWorks for Beginners - Udemy",               "platform": "https://www.udemy.com",                 "project": "Design a gear assembly",                       "duration": "3 weeks"},
    "ANSYS":              {"course": "ANSYS for Beginners - Udemy",                    "platform": "https://www.udemy.com",                 "project": "Perform stress analysis on a component",       "duration": "3 weeks"},
    "STAAD Pro":          {"course": "STAAD Pro Complete Course - Udemy",              "platform": "https://www.udemy.com",                 "project": "Design a residential building structure",      "duration": "3 weeks"},
    "PLC Programming":    {"course": "PLC Programming for Beginners - Udemy",          "platform": "https://www.udemy.com",                 "project": "Automate a conveyor belt system using PLC",    "duration": "3 weeks"},
    "VLSI":               {"course": "VLSI Design - NPTEL",                            "platform": "https://nptel.ac.in",                   "project": "Design a basic ALU using Verilog",             "duration": "4 weeks"},
    "Embedded C":         {"course": "Embedded C Programming - Udemy",                 "platform": "https://www.udemy.com",                 "project": "Build a temperature monitor using Arduino",    "duration": "3 weeks"},
    "Verilog":            {"course": "Verilog HDL - Udemy",                            "platform": "https://www.udemy.com",                 "project": "Implement a 4-bit counter in Verilog",         "duration": "3 weeks"},
    "General Studies - History": {"course": "History for UPSC - YouTube (Unacademy)", "platform": "https://www.youtube.com",               "project": "Make chapter-wise notes and revise weekly",    "duration": "8 weeks"},
    "General Studies - Polity & Constitution": {"course": "Indian Polity by M Laxmikanth", "platform": "https://www.amazon.in",             "project": "Solve previous year polity questions daily",   "duration": "6 weeks"},
    "General Studies - Indian Economy": {"course": "Indian Economy by Ramesh Singh",  "platform": "https://www.amazon.in",                 "project": "Write 5 economy answer practice sheets",       "duration": "6 weeks"},
    "Current Affairs (The Hindu / Indian Express)": {"course": "Daily Newspaper Reading + Vision IAS Monthly", "platform": "https://www.visionias.in", "project": "Maintain a current affairs diary monthly", "duration": "Ongoing"},
    "CSAT - Aptitude & Reasoning": {"course": "CSAT Paper 2 - Arihant Publications", "platform": "https://www.amazon.in",                  "project": "Solve 20 CSAT questions daily",                "duration": "4 weeks"},
    "Essay Writing":      {"course": "Essay Writing for UPSC - YouTube (Insights IAS)","platform": "https://www.youtube.com",              "project": "Write 2 essays per week and get feedback",     "duration": "4 weeks"},
    "Quantitative Aptitude - Number System & DI": {"course": "Quantitative Aptitude by R.S. Aggarwal", "platform": "https://www.amazon.in", "project": "Solve 30 DI sets from previous papers",      "duration": "4 weeks"},
    "Reasoning Ability - Puzzles & Seating Arrangement": {"course": "Logical Reasoning - Oliveboard App", "platform": "https://www.oliveboard.in", "project": "Solve 10 puzzles daily",                "duration": "3 weeks"},
    "General Awareness - Banking & Economy": {"course": "Banking Awareness by Arihant", "platform": "https://www.amazon.in",               "project": "Revise 50 banking terms and RBI policies",    "duration": "2 weeks"},
    "Mock Test Practice": {"course": "Mock Tests - Testbook / Oliveboard",             "platform": "https://www.testbook.com",              "project": "Take 1 full mock test every 3 days",           "duration": "Ongoing"},
    "Mathematics - Arithmetic & Algebra": {"course": "Maths for Competitive Exams - Unacademy", "platform": "https://www.youtube.com",      "project": "Solve 25 arithmetic questions daily",          "duration": "4 weeks"},
    "Kannada Language & Grammar": {"course": "Kannada Grammar for KPSC - YouTube",    "platform": "https://www.youtube.com",               "project": "Write 2 Kannada essays per week",              "duration": "4 weeks"},
    "Karnataka History & Culture": {"course": "Karnataka GK - Unique Publications",   "platform": "https://www.amazon.in",                 "project": "Make district-wise Karnataka notes",           "duration": "4 weeks"},
    # ── BSc Forensic Science ──────────────────────────────────────
"Crime Scene Investigation":              {"course": "Introduction to Forensic Science - Coursera (Nanyang Tech)",   "platform": "https://www.coursera.org",                       "project": "Document and reconstruct a mock crime scene with photos and notes",   "duration": "4 weeks"},
"Forensic Photography":                   {"course": "Forensic Photography Basics - Udemy",                          "platform": "https://www.udemy.com",                          "project": "Create a portfolio of mock evidence photographs with scale markers",   "duration": "2 weeks"},
"DNA Analysis":                           {"course": "Genomic Data Science - Coursera (Johns Hopkins)",              "platform": "https://www.coursera.org",                       "project": "Analyse a sample DNA sequence and write a forensic report",            "duration": "6 weeks"},
"Fingerprint Analysis":                   {"course": "Fingerprint Identification - YouTube (Forensic Files)",        "platform": "https://www.youtube.com",                        "project": "Collect and classify 20 fingerprint samples using arch/loop/whorl",    "duration": "3 weeks"},
"Toxicology":                             {"course": "Forensic Toxicology - NPTEL",                                  "platform": "https://nptel.ac.in",                            "project": "Prepare a case study on a real toxicology poisoning case",             "duration": "4 weeks"},
"Ballistics & Firearm Analysis":          {"course": "Forensic Ballistics - YouTube (Crash Course Forensics)",       "platform": "https://www.youtube.com",                        "project": "Study and document 10 bullet trajectory examples",                     "duration": "3 weeks"},
"Forensic Chemistry":                     {"course": "Forensic Chemistry & Analysis - NPTEL",                        "platform": "https://nptel.ac.in",                            "project": "Perform a mock drug identification using chemical tests",              "duration": "4 weeks"},
"Forensic Biology":                       {"course": "Forensic Biology Fundamentals - NPTEL",                        "platform": "https://nptel.ac.in",                            "project": "Analyse biological samples (hair, fibre, blood) under microscope",     "duration": "4 weeks"},
"Indian Evidence Act":                    {"course": "Indian Evidence Act 1872 - Unacademy Legal",                   "platform": "https://unacademy.com",                          "project": "Write case briefs on 5 landmark evidence-related judgments",           "duration": "3 weeks"},
"Criminal Law (IPC, CrPC, BNS)":          {"course": "Criminal Law Masterclass - LawSikho",                          "platform": "https://lawsikho.com",                           "project": "Compare key provisions of old IPC vs new BNS",                         "duration": "5 weeks"},
"Microscopy Techniques":                  {"course": "Microscopy for Forensic Sciences - Coursera",                  "platform": "https://www.coursera.org",                       "project": "Identify and photograph 15 trace evidence samples",                    "duration": "3 weeks"},
"Document Examination":                   {"course": "Questioned Document Examination - Udemy",                      "platform": "https://www.udemy.com",                          "project": "Compare and analyse handwriting samples from 5 sources",               "duration": "3 weeks"},
"Cyber Forensics":                        {"course": "Computer Forensics Fundamentals - NPTEL",                      "platform": "https://nptel.ac.in",                            "project": "Perform a mock digital evidence recovery using FTK Imager or Autopsy", "duration": "5 weeks"},
"Report Writing":                         {"course": "Technical & Forensic Report Writing - Udemy",                  "platform": "https://www.udemy.com",                          "project": "Write a complete forensic case report for a mock investigation",       "duration": "2 weeks"},
"Chain of Custody Procedures":            {"course": "Chain of Custody in Forensics - YouTube (NIJ Training)",       "platform": "https://www.youtube.com",                        "project": "Design a mock evidence chain-of-custody log for 10 items",             "duration": "2 weeks"},

# ── Biology / Microbiology / Biotech ──────────────────────────
"Microbiology":                           {"course": "Microbiology Fundamentals - NPTEL",                            "platform": "https://nptel.ac.in",                            "project": "Identify and document 10 common microbes in a lab notebook",           "duration": "4 weeks"},
"Cell Biology":                           {"course": "Cell Biology - Coursera (UC San Diego)",                       "platform": "https://www.coursera.org",                       "project": "Prepare a cell organelle reference chart with functions",              "duration": "4 weeks"},
"Genetics":                               {"course": "Introduction to Genetics - Coursera (Duke University)",        "platform": "https://www.coursera.org",                       "project": "Solve 30 Mendelian inheritance problems with Punnett squares",         "duration": "4 weeks"},
"Biochemistry":                           {"course": "Biochemistry Basics - NPTEL",                                  "platform": "https://nptel.ac.in",                            "project": "Map the glycolysis pathway and write a detailed report",               "duration": "5 weeks"},
"Anatomy & Physiology":                   {"course": "Anatomy & Physiology - Coursera (Univ of Michigan)",           "platform": "https://www.coursera.org",                       "project": "Prepare flash cards for 50 anatomical structures",                     "duration": "6 weeks"},
"Ecology":                                {"course": "Ecology - NPTEL",                                              "platform": "https://nptel.ac.in",                            "project": "Conduct a local biodiversity survey of 1 km radius",                   "duration": "3 weeks"},
"Molecular Biology":                      {"course": "Molecular Biology - Coursera (MIT)",                           "platform": "https://www.coursera.org",                       "project": "Build a 3D model of DNA replication using cardboard",                  "duration": "5 weeks"},
"Laboratory Techniques":                  {"course": "Laboratory Techniques - YouTube (Khan Academy)",               "platform": "https://www.youtube.com",                        "project": "Maintain a 20-experiment lab notebook with results",                   "duration": "Ongoing"},
"Scientific Writing":                     {"course": "Writing in the Sciences - Coursera (Stanford)",                "platform": "https://www.coursera.org",                       "project": "Write a 5-page research paper in IEEE format",                         "duration": "3 weeks"},
"Microbial Cultivation Techniques":       {"course": "Microbiological Techniques - NPTEL",                           "platform": "https://nptel.ac.in",                            "project": "Document 10 culture media preparations",                               "duration": "3 weeks"},
"Sterilization Methods":                  {"course": "Sterilization Techniques - YouTube (BioBox)",                  "platform": "https://www.youtube.com",                        "project": "Create a comparison chart of 5 sterilization methods",                 "duration": "1 week"},
"Microbial Identification":               {"course": "Bacterial Identification - Coursera (Johns Hopkins)",          "platform": "https://www.coursera.org",                       "project": "Identify 10 bacterial isolates using biochemical tests",               "duration": "4 weeks"},
"PCR Techniques":                         {"course": "PCR Techniques - Coursera (UCSD)",                             "platform": "https://www.coursera.org",                       "project": "Design primers for a gene of your choice",                             "duration": "3 weeks"},
"PCR & Gel Electrophoresis":              {"course": "Molecular Biology Techniques - NPTEL",                         "platform": "https://nptel.ac.in",                            "project": "Simulate gel electrophoresis results for 5 DNA samples",               "duration": "3 weeks"},
"Microbial Genetics":                     {"course": "Microbial Genetics - NPTEL",                                   "platform": "https://nptel.ac.in",                            "project": "Trace gene transfer in 3 microbial cases",                             "duration": "4 weeks"},
"Immunology":                             {"course": "Fundamentals of Immunology - Coursera (Rice University)",      "platform": "https://www.coursera.org",                       "project": "Draw and explain the antibody-antigen interaction",                    "duration": "5 weeks"},
"Bioinformatics":                         {"course": "Bioinformatics Specialization - Coursera (UC San Diego)",      "platform": "https://www.coursera.org",                       "project": "Perform BLAST search on 5 sequences using NCBI",                       "duration": "6 weeks"},
"Genetic Engineering":                    {"course": "Genetic Engineering - NPTEL",                                  "platform": "https://nptel.ac.in",                            "project": "Design a recombinant DNA cloning experiment on paper",                 "duration": "5 weeks"},
"Cell Culture Techniques":                {"course": "Cell Culture Methods - YouTube (iBiology)",                    "platform": "https://www.youtube.com",                        "project": "Document protocols for 3 cell culture techniques",                     "duration": "3 weeks"},
"Fermentation Technology":                {"course": "Fermentation Technology - NPTEL",                              "platform": "https://nptel.ac.in",                            "project": "Design a small-scale fermentation process for a product",              "duration": "4 weeks"},

# ── BSc Chemistry skills ──────────────────────────────────────
"Organic Chemistry":                      {"course": "Organic Chemistry I - Coursera (Univ of Manchester)",          "platform": "https://www.coursera.org",                       "project": "Solve and document 30 reaction mechanism problems",                    "duration": "6 weeks"},
"Inorganic Chemistry":                    {"course": "Inorganic Chemistry - NPTEL",                                  "platform": "https://nptel.ac.in",                            "project": "Build models of 10 coordination compounds",                            "duration": "5 weeks"},
"Physical Chemistry":                     {"course": "Physical Chemistry - NPTEL",                                   "platform": "https://nptel.ac.in",                            "project": "Solve 25 thermodynamics and kinetics problems",                        "duration": "5 weeks"},
"Analytical Chemistry":                   {"course": "Analytical Chemistry - Coursera (Univ of Tartu)",              "platform": "https://www.coursera.org",                       "project": "Perform 5 titration experiments and document results",                 "duration": "4 weeks"},
"Chromatography (HPLC / GC)":             {"course": "Chromatography Techniques - YouTube (Royce Murray)",           "platform": "https://www.youtube.com",                        "project": "Analyse a chromatogram and interpret peaks for 3 samples",             "duration": "3 weeks"},
"Spectroscopy":                           {"course": "Spectroscopy in Chemistry - NPTEL",                            "platform": "https://nptel.ac.in",                            "project": "Identify an unknown compound from IR/NMR/MS data",                     "duration": "4 weeks"},
"Quality Control":                        {"course": "Quality Control & Assurance - Coursera",                       "platform": "https://www.coursera.org",                       "project": "Design a QC checklist for a pharmaceutical product",                   "duration": "3 weeks"},

# ── BSc Physics skills ────────────────────────────────────────
"Classical Mechanics":                    {"course": "Classical Mechanics - MIT OpenCourseware",                     "platform": "https://ocw.mit.edu",                            "project": "Solve 25 Newtonian mechanics problems with diagrams",                   "duration": "6 weeks"},
"Quantum Mechanics":                      {"course": "Quantum Mechanics - NPTEL",                                    "platform": "https://nptel.ac.in",                            "project": "Solve the particle-in-a-box problem with energy diagrams",             "duration": "8 weeks"},
"Electromagnetism":                       {"course": "Electromagnetism - MIT OpenCourseware",                        "platform": "https://ocw.mit.edu",                            "project": "Build a simple electromagnet and document field strength",             "duration": "6 weeks"},
"Optics":                                 {"course": "Introduction to Optics - Coursera (Univ of Colorado)",         "platform": "https://www.coursera.org",                       "project": "Design a simple telescope or microscope on paper",                     "duration": "4 weeks"},
"Mathematical Physics":                   {"course": "Mathematical Methods for Physics - NPTEL",                     "platform": "https://nptel.ac.in",                            "project": "Solve 20 problems on Fourier series and PDEs",                         "duration": "6 weeks"},
"Scientific Communication":               {"course": "Communicating Science Effectively - Coursera",                 "platform": "https://www.coursera.org",                       "project": "Present a science topic in 3 minutes for a general audience",          "duration": "2 weeks"},

# ── Math / Statistics ────────────────────────────────────────
"Statistics":                             {"course": "Introduction to Statistics - Coursera (Stanford)",             "platform": "https://www.coursera.org",                       "project": "Analyse a real dataset and present descriptive statistics",            "duration": "5 weeks"},
"Probability":                            {"course": "Probability - Coursera (MIT)",                                 "platform": "https://www.coursera.org",                       "project": "Build a simulation of 5 classic probability problems",                 "duration": "5 weeks"},
"Linear Algebra":                         {"course": "Linear Algebra - Coursera (Imperial College)",                 "platform": "https://www.coursera.org",                       "project": "Solve 20 matrix operations and eigenvalue problems",                   "duration": "5 weeks"},
"Calculus":                               {"course": "Calculus - Khan Academy",                                      "platform": "https://www.khanacademy.org",                    "project": "Solve 30 differentiation and integration problems",                    "duration": "6 weeks"},
"Discrete Mathematics":                   {"course": "Discrete Math - Coursera (UC San Diego)",                      "platform": "https://www.coursera.org",                       "project": "Solve 25 graph theory and logic problems",                             "duration": "5 weeks"},
"R Programming":                          {"course": "R Programming - Coursera (Johns Hopkins)",                     "platform": "https://www.coursera.org",                       "project": "Perform exploratory data analysis on a dataset in R",                  "duration": "4 weeks"},
"SPSS":                                   {"course": "SPSS for Beginners - Udemy",                                   "platform": "https://www.udemy.com",                          "project": "Conduct a t-test and ANOVA on a survey dataset",                       "duration": "3 weeks"},
"Data Analysis":                          {"course": "Data Analysis - Coursera (Johns Hopkins)",                     "platform": "https://www.coursera.org",                       "project": "Clean and analyse a public dataset, write a report",                   "duration": "4 weeks"},
"NumPy":                                  {"course": "NumPy Tutorial - Kaggle Learn",                                "platform": "https://www.kaggle.com/learn",                   "project": "Solve 25 NumPy array manipulation exercises",                          "duration": "2 weeks"},
"Data Cleaning":                          {"course": "Data Cleaning - Kaggle Learn",                                 "platform": "https://www.kaggle.com/learn",                   "project": "Clean and document a real messy dataset",                              "duration": "2 weeks"},
"Big Data Basics":                        {"course": "Big Data Specialization - Coursera (UC San Diego)",            "platform": "https://www.coursera.org",                       "project": "Process a small dataset using Hadoop or Spark",                        "duration": "6 weeks"},

# ── BSc Nursing ──────────────────────────────────────────────
"Patient Care":                           {"course": "Patient Care Skills - Coursera (Stanford Medicine)",           "platform": "https://www.coursera.org",                       "project": "Document 10 patient care scenarios with empathy notes",                "duration": "4 weeks"},
"Pharmacology":                           {"course": "Pharmacology - Coursera (UC San Diego)",                       "platform": "https://www.coursera.org",                       "project": "Compile a drug reference chart for 25 common medications",             "duration": "6 weeks"},
"Medical-Surgical Nursing":               {"course": "Medical-Surgical Nursing - NPTEL",                             "platform": "https://nptel.ac.in",                            "project": "Write care plans for 5 common surgical conditions",                    "duration": "8 weeks"},
"First Aid & Emergency Care":             {"course": "First Aid Certification - Red Cross",                          "platform": "https://www.redcross.org",                       "project": "Demonstrate CPR and basic first aid for 10 scenarios",                 "duration": "2 weeks"},
"Infection Control":                      {"course": "Infection Prevention - Coursera (Johns Hopkins)",              "platform": "https://www.coursera.org",                       "project": "Create an infection control SOP for a small clinic",                   "duration": "3 weeks"},
"Vital Signs Monitoring":                 {"course": "Vital Signs Monitoring - YouTube (NurseStudy)",                "platform": "https://www.youtube.com",                        "project": "Practise measuring all vital signs on 10 mock patients",               "duration": "1 week"},
"Medical Documentation":                  {"course": "Health Record Documentation - Coursera",                       "platform": "https://www.coursera.org",                       "project": "Maintain a 20-patient mock chart with progress notes",                 "duration": "2 weeks"},
"Empathy & Counseling":                   {"course": "Empathy in Healthcare - Coursera (Stanford)",                  "platform": "https://www.coursera.org",                       "project": "Role-play 5 patient counseling scenarios",                             "duration": "3 weeks"},
"Nutrition":                              {"course": "Nutrition and Health - Coursera (Stanford)",                   "platform": "https://www.coursera.org",                       "project": "Design a 7-day diet plan for 3 patient profiles",                      "duration": "3 weeks"},
"Community Health":                       {"course": "Community Health Nursing - NPTEL",                             "platform": "https://nptel.ac.in",                            "project": "Conduct a mock community health survey of 20 households",              "duration": "4 weeks"},
"Hospital Information Systems":           {"course": "Healthcare IT - Coursera (Johns Hopkins)",                     "platform": "https://www.coursera.org",                       "project": "Document workflows for 5 hospital department modules",                 "duration": "3 weeks"},

# ── BSc Agriculture ──────────────────────────────────────────
"Soil Science":                           {"course": "Soil Science Fundamentals - NPTEL",                            "platform": "https://nptel.ac.in",                            "project": "Test 5 soil samples for pH, N, P, K levels",                           "duration": "4 weeks"},
"Crop Production":                        {"course": "Principles of Crop Production - NPTEL",                        "platform": "https://nptel.ac.in",                            "project": "Plan a crop cycle for 1 acre over 12 months",                          "duration": "5 weeks"},
"Plant Pathology":                        {"course": "Plant Pathology - NPTEL",                                      "platform": "https://nptel.ac.in",                            "project": "Identify and document 10 common plant diseases",                       "duration": "4 weeks"},
"Horticulture":                           {"course": "Horticulture - NPTEL",                                         "platform": "https://nptel.ac.in",                            "project": "Maintain a small kitchen garden for 8 weeks",                          "duration": "6 weeks"},
"Agronomy":                               {"course": "Agronomy Fundamentals - NPTEL",                                "platform": "https://nptel.ac.in",                            "project": "Compare yield of 3 crops in your local region",                        "duration": "5 weeks"},
"Irrigation Management":                  {"course": "Irrigation Engineering - NPTEL",                               "platform": "https://nptel.ac.in",                            "project": "Design a drip irrigation layout for 1 acre",                           "duration": "3 weeks"},
"Agricultural Economics":                 {"course": "Agricultural Economics - Coursera",                            "platform": "https://www.coursera.org",                       "project": "Analyse cost vs. revenue for a sample farm",                           "duration": "4 weeks"},
"Pest Management":                        {"course": "Integrated Pest Management - NPTEL",                           "platform": "https://nptel.ac.in",                            "project": "Develop an IPM plan for a real or hypothetical farm",                  "duration": "3 weeks"},
"Agribusiness":                           {"course": "Agribusiness Management - Coursera",                           "platform": "https://www.coursera.org",                       "project": "Create a business plan for a small agri-startup",                      "duration": "4 weeks"},
"GIS for Agriculture":                    {"course": "GIS for Agriculture - Coursera (UC Davis)",                    "platform": "https://www.coursera.org",                       "project": "Map crop distribution of your district using QGIS",                    "duration": "4 weeks"},
"Field Survey Techniques":                {"course": "Agricultural Survey Methods - YouTube (ICAR)",                 "platform": "https://www.youtube.com",                        "project": "Conduct a small farm survey using GPS tools",                          "duration": "2 weeks"},

# ── BSc Psychology ───────────────────────────────────────────
"Counselling Techniques":                 {"course": "Counselling Skills - Coursera (Univ of Toronto)",              "platform": "https://www.coursera.org",                       "project": "Conduct 3 mock counselling sessions with reflection notes",            "duration": "5 weeks"},
"Cognitive Psychology":                   {"course": "Cognitive Psychology - Coursera (Yale)",                       "platform": "https://www.coursera.org",                       "project": "Design 3 cognitive bias experiments with 10 subjects",                 "duration": "5 weeks"},
"Developmental Psychology":               {"course": "Child Development - Coursera (Univ of California)",            "platform": "https://www.coursera.org",                       "project": "Observe and document a child's behaviour across 4 stages",             "duration": "5 weeks"},
"Abnormal Psychology":                    {"course": "Abnormal Psychology - NPTEL",                                  "platform": "https://nptel.ac.in",                            "project": "Write case studies on 5 disorders from DSM-5",                         "duration": "6 weeks"},
"Social Psychology":                      {"course": "Social Psychology - Coursera (Wesleyan University)",           "platform": "https://www.coursera.org",                       "project": "Conduct a small social experiment with 10 participants",               "duration": "5 weeks"},
"Psychological Assessment":               {"course": "Psychological Assessment - NPTEL",                             "platform": "https://nptel.ac.in",                            "project": "Administer and interpret 3 standard psychological tests",              "duration": "5 weeks"},
"Empathy & Active Listening":             {"course": "Active Listening - Coursera (UC Davis)",                       "platform": "https://www.coursera.org",                       "project": "Practise active listening in 10 real conversations",                   "duration": "3 weeks"},
"Case Study Writing":                     {"course": "Case Study Method - Coursera (Harvard Business)",              "platform": "https://www.coursera.org",                       "project": "Write 3 well-structured psychological case studies",                   "duration": "3 weeks"},

# ── BPharm ───────────────────────────────────────────────────
"Pharmaceutics":                          {"course": "Pharmaceutics - NPTEL",                                        "platform": "https://nptel.ac.in",                            "project": "Formulate and document 3 dosage forms (tablet/syrup/cream)",           "duration": "5 weeks"},
"Pharmacognosy":                          {"course": "Pharmacognosy - NPTEL",                                        "platform": "https://nptel.ac.in",                            "project": "Identify and document 10 medicinal plants and their uses",             "duration": "4 weeks"},
"Medicinal Chemistry":                    {"course": "Medicinal Chemistry - NPTEL",                                  "platform": "https://nptel.ac.in",                            "project": "Study structure-activity relationships of 5 drug classes",             "duration": "6 weeks"},
"Pharmaceutical Analysis":                {"course": "Pharmaceutical Analysis - NPTEL",                              "platform": "https://nptel.ac.in",                            "project": "Perform titrations and assay tests on 5 mock pharma samples",          "duration": "5 weeks"},
"Clinical Pharmacy":                      {"course": "Clinical Pharmacy - Coursera (Univ of Sydney)",                "platform": "https://www.coursera.org",                       "project": "Review medication therapy for 5 sample patient profiles",              "duration": "5 weeks"},
"Hospital Pharmacy":                      {"course": "Hospital Pharmacy Practice - YouTube (PharmaGuide)",           "platform": "https://www.youtube.com",                        "project": "Design a hospital pharmacy workflow for a 100-bed hospital",           "duration": "3 weeks"},
"Patient Counseling":                     {"course": "Pharmacist Patient Counseling - Coursera",                     "platform": "https://www.coursera.org",                       "project": "Role-play 10 patient counseling scenarios for common drugs",           "duration": "3 weeks"},
"Drug Safety":                            {"course": "Pharmacovigilance - Coursera (Univ of Geneva)",                "platform": "https://www.coursera.org",                       "project": "Document 5 ADR cases following pharmacovigilance guidelines",          "duration": "4 weeks"},
"Quality Assurance":                      {"course": "Pharmaceutical QA & GMP - Udemy",                              "platform": "https://www.udemy.com",                          "project": "Create a GMP audit checklist for a pharma facility",                   "duration": "4 weeks"},
"Pharmaceutical Marketing":               {"course": "Pharma Marketing - Coursera (Yale School of Mgmt)",            "platform": "https://www.coursera.org",                       "project": "Build a marketing plan for a hypothetical new drug",                   "duration": "3 weeks"},
"Drug Regulatory Affairs":                {"course": "Drug Regulatory Affairs - NPTEL",                              "platform": "https://nptel.ac.in",                            "project": "Compile regulatory dossiers for 3 mock drug applications",             "duration": "4 weeks"},

# ── BHM (Hotel Management) ───────────────────────────────────
"Food Production":                        {"course": "Culinary Arts Fundamentals - Coursera (Le Cordon Bleu)",       "platform": "https://www.coursera.org",                       "project": "Prepare and document 10 dishes from international cuisines",           "duration": "6 weeks"},
"F&B Service":                            {"course": "Food & Beverage Service - YouTube (HotelManagementStudies)",   "platform": "https://www.youtube.com",                        "project": "Practise service standards for 5 dining-room scenarios",               "duration": "3 weeks"},
"Front Office Management":                {"course": "Hotel Front Office Operations - Udemy",                        "platform": "https://www.udemy.com",                          "project": "Simulate check-in/check-out for 20 guests with software",              "duration": "3 weeks"},
"Housekeeping Management":                {"course": "Housekeeping Operations - Udemy",                              "platform": "https://www.udemy.com",                          "project": "Design SOPs for room cleaning and laundry operations",                 "duration": "3 weeks"},
"Hotel Accounts":                         {"course": "Hospitality Accounting - Coursera (ESSEC)",                    "platform": "https://www.coursera.org",                       "project": "Maintain a month's hotel accounts in Excel",                           "duration": "3 weeks"},
"Tourism Management":                     {"course": "Tourism Management - Coursera",                                "platform": "https://www.coursera.org",                       "project": "Design a 7-day tour package for a popular destination",                "duration": "3 weeks"},
"Event Management":                       {"course": "Event Management - Udemy",                                     "platform": "https://www.udemy.com",                          "project": "Plan a 200-guest wedding event end-to-end",                            "duration": "4 weeks"},
"Customer Service":                       {"course": "Customer Service Skills - Coursera (Univ of Penn)",            "platform": "https://www.coursera.org",                       "project": "Document handling of 10 difficult customer scenarios",                 "duration": "3 weeks"},
"Culinary Arts":                          {"course": "Culinary Arts - Coursera (Univ of California)",                "platform": "https://www.coursera.org",                       "project": "Create a 5-course tasting menu with recipes",                          "duration": "5 weeks"},
"Beverage Knowledge":                     {"course": "Bartending Fundamentals - Udemy",                              "platform": "https://www.udemy.com",                          "project": "Memorise and prepare 20 classic cocktails",                            "duration": "3 weeks"},
"Hospitality Marketing":                  {"course": "Hospitality Marketing - Coursera (ESSEC)",                     "platform": "https://www.coursera.org",                       "project": "Build a social media marketing plan for a small hotel",                "duration": "3 weeks"},
"Hospitality Law":                        {"course": "Hospitality Law - YouTube (Hospitality Now)",                  "platform": "https://www.youtube.com",                        "project": "Summarise 10 common hotel/guest legal disputes",                       "duration": "2 weeks"},

# ── MECHANICAL ENGINEERING — fill in the gaps ─────────────────
"CATIA":                                  {"course": "CATIA V5 Complete Course - Udemy",                             "platform": "https://www.udemy.com",                          "project": "Model and assemble a small mechanical component",                      "duration": "4 weeks"},
"CFD":                                    {"course": "Computational Fluid Dynamics - NPTEL",                         "platform": "https://nptel.ac.in",                            "project": "Simulate airflow over a 2D airfoil using OpenFOAM or ANSYS",           "duration": "5 weeks"},
"Thermodynamics":                         {"course": "Thermodynamics - MIT OpenCourseware",                          "platform": "https://ocw.mit.edu",                            "project": "Solve 25 thermodynamic cycle problems with PV diagrams",               "duration": "6 weeks"},
"Fluid Mechanics":                        {"course": "Fluid Mechanics - NPTEL",                                      "platform": "https://nptel.ac.in",                            "project": "Simulate fluid flow in a pipe using ANSYS Fluent",                     "duration": "5 weeks"},
"Heat Transfer":                          {"course": "Heat Transfer - NPTEL",                                        "platform": "https://nptel.ac.in",                            "project": "Design a heat exchanger for industrial use",                           "duration": "5 weeks"},
"Manufacturing Processes":                {"course": "Manufacturing Processes - NPTEL",                              "platform": "https://nptel.ac.in",                            "project": "Document 10 metal-forming and machining processes",                    "duration": "5 weeks"},
"GD&T":                                   {"course": "GD&T Fundamentals - Udemy",                                    "platform": "https://www.udemy.com",                          "project": "Annotate 10 engineering drawings with proper GD&T symbols",            "duration": "3 weeks"},
"Finite Element Analysis":                {"course": "FEA - NPTEL",                                                  "platform": "https://nptel.ac.in",                            "project": "Perform FEA on a bracket and analyse stress concentrations",           "duration": "5 weeks"},
"Lean Manufacturing":                     {"course": "Lean Production - Coursera (Univ of Michigan)",                "platform": "https://www.coursera.org",                       "project": "Apply 5S and Kaizen to a small workshop",                              "duration": "4 weeks"},
"Six Sigma":                              {"course": "Six Sigma Yellow Belt - Coursera (Univ of Georgia)",           "platform": "https://www.coursera.org",                       "project": "Solve a mock manufacturing problem using DMAIC",                       "duration": "5 weeks"},
"Project Management":                     {"course": "Google Project Management Certificate - Coursera",             "platform": "https://www.coursera.org",                       "project": "Plan a 3-month project with Gantt chart and milestones",               "duration": "5 weeks"},

# ── ECE / EEE — fill in the gaps ─────────────────────────────
"Embedded C":                             {"course": "Embedded C Programming - Udemy",                               "platform": "https://www.udemy.com",                          "project": "Build a temperature monitor using Arduino",                            "duration": "3 weeks"},
"Verilog":                                {"course": "Verilog HDL - NPTEL",                                          "platform": "https://nptel.ac.in",                            "project": "Implement a 4-bit counter and 4-bit ALU in Verilog",                   "duration": "3 weeks"},
"VLSI":                                   {"course": "VLSI Design - NPTEL",                                          "platform": "https://nptel.ac.in",                            "project": "Design a basic ALU in Verilog and simulate it",                        "duration": "4 weeks"},
"PCB Design":                             {"course": "PCB Design with KiCad - Udemy",                                "platform": "https://www.udemy.com",                          "project": "Design and route a small electronic circuit PCB",                      "duration": "3 weeks"},
"Signal Processing":                      {"course": "Digital Signal Processing - Coursera (EPFL)",                  "platform": "https://www.coursera.org",                       "project": "Filter audio signals using MATLAB or Python",                          "duration": "5 weeks"},
"Microcontrollers":                       {"course": "Microcontroller Programming - NPTEL",                          "platform": "https://nptel.ac.in",                            "project": "Build 3 microcontroller-based projects (LED, sensor, motor)",          "duration": "4 weeks"},
"Arduino":                                {"course": "Arduino Step by Step - Udemy",                                 "platform": "https://www.udemy.com",                          "project": "Build a smart-home prototype using Arduino",                           "duration": "3 weeks"},
"Raspberry Pi":                           {"course": "Raspberry Pi for Beginners - Udemy",                           "platform": "https://www.udemy.com",                          "project": "Build a home automation system with Raspberry Pi",                     "duration": "4 weeks"},
"Communication Systems":                  {"course": "Principles of Communications - NPTEL",                         "platform": "https://nptel.ac.in",                            "project": "Simulate AM/FM modulation in MATLAB or Python",                        "duration": "5 weeks"},
"Circuit Design":                         {"course": "Circuit Analysis - Coursera (Georgia Tech)",                   "platform": "https://www.coursera.org",                       "project": "Design and simulate 5 analog circuits in LTspice",                     "duration": "4 weeks"},
"LabVIEW":                                {"course": "LabVIEW for Beginners - Udemy",                                "platform": "https://www.udemy.com",                          "project": "Build a virtual instrument for data acquisition",                      "duration": "3 weeks"},
"Antenna Design":                         {"course": "Antenna Theory - NPTEL",                                       "platform": "https://nptel.ac.in",                            "project": "Design and simulate a dipole or patch antenna",                        "duration": "4 weeks"},
"IoT":                                    {"course": "Internet of Things - Coursera (UC Irvine)",                    "platform": "https://www.coursera.org",                       "project": "Build an IoT-enabled sensor node with cloud reporting",                "duration": "5 weeks"},
"PLC Programming":                        {"course": "PLC Programming for Beginners - Udemy",                        "platform": "https://www.udemy.com",                          "project": "Automate a conveyor belt using PLC ladder logic",                      "duration": "3 weeks"},
"SCADA":                                  {"course": "SCADA Fundamentals - Udemy",                                   "platform": "https://www.udemy.com",                          "project": "Design a SCADA HMI for a water-treatment plant",                       "duration": "3 weeks"},
"AutoCAD Electrical":                     {"course": "AutoCAD Electrical Complete - Udemy",                          "platform": "https://www.udemy.com",                          "project": "Create electrical schematics for a small panel",                       "duration": "3 weeks"},
"Power Systems":                          {"course": "Power Systems Engineering - NPTEL",                            "platform": "https://nptel.ac.in",                            "project": "Analyse load flow on a small 3-bus power network",                     "duration": "5 weeks"},
"Control Systems":                        {"course": "Control Systems - NPTEL",                                      "platform": "https://nptel.ac.in",                            "project": "Design a PID controller and simulate it in MATLAB",                    "duration": "5 weeks"},
"Electrical Machines":                    {"course": "Electrical Machines - NPTEL",                                  "platform": "https://nptel.ac.in",                            "project": "Build a working model of a DC motor or transformer",                   "duration": "4 weeks"},
"Power Electronics":                      {"course": "Power Electronics - NPTEL",                                    "platform": "https://nptel.ac.in",                            "project": "Design a simple converter circuit (buck/boost)",                       "duration": "5 weeks"},
"Switchgear":                             {"course": "Switchgear & Protection - YouTube (NEEPCO)",                   "platform": "https://www.youtube.com",                        "project": "Document working of 5 switchgear types with diagrams",                 "duration": "2 weeks"},
"Solar PV Systems":                       {"course": "Solar PV Systems - Coursera (SUNY)",                           "platform": "https://www.coursera.org",                       "project": "Design a 5kW rooftop solar installation on paper",                     "duration": "4 weeks"},
"Protection Systems":                     {"course": "Power System Protection - NPTEL",                              "platform": "https://nptel.ac.in",                            "project": "Design protection relay settings for a small distribution feeder",     "duration": "4 weeks"},

# ── CIVIL — fill in the gaps ─────────────────────────────────
"ETABS":                                  {"course": "ETABS for Structural Analysis - Udemy",                        "platform": "https://www.udemy.com",                          "project": "Design a 5-storey RC building model in ETABS",                         "duration": "4 weeks"},
"STAAD Pro":                              {"course": "STAAD Pro Complete - Udemy",                                   "platform": "https://www.udemy.com",                          "project": "Design a residential building structure",                              "duration": "3 weeks"},
"Revit":                                  {"course": "Revit Architecture - Udemy",                                   "platform": "https://www.udemy.com",                          "project": "Model a complete 2-storey villa in Revit",                             "duration": "4 weeks"},
"MS Project":                             {"course": "Microsoft Project Essentials - Udemy",                         "platform": "https://www.udemy.com",                          "project": "Schedule a 6-month construction project with Gantt charts",            "duration": "3 weeks"},
"Structural Analysis":                    {"course": "Structural Analysis - NPTEL",                                  "platform": "https://nptel.ac.in",                            "project": "Solve 20 problems on beams, trusses and frames",                       "duration": "5 weeks"},
"Surveying":                              {"course": "Surveying - NPTEL",                                            "platform": "https://nptel.ac.in",                            "project": "Conduct a chain or theodolite survey of a small plot",                 "duration": "3 weeks"},
"Estimation & Costing":                   {"course": "Construction Estimating - Udemy",                              "platform": "https://www.udemy.com",                          "project": "Prepare a BOQ for a small residential building",                       "duration": "4 weeks"},
"Concrete Technology":                    {"course": "Concrete Technology - NPTEL",                                  "platform": "https://nptel.ac.in",                            "project": "Design mix proportions for 3 grades of concrete",                      "duration": "4 weeks"},
"Soil Mechanics":                         {"course": "Soil Mechanics - NPTEL",                                       "platform": "https://nptel.ac.in",                            "project": "Test and classify 3 local soil samples",                               "duration": "4 weeks"},
"GIS":                                    {"course": "GIS Fundamentals - Coursera (UC Davis)",                       "platform": "https://www.coursera.org",                       "project": "Create a digital map of your locality using QGIS",                     "duration": "4 weeks"},
"Construction Management":                {"course": "Construction Management - Coursera (Columbia)",                "platform": "https://www.coursera.org",                       "project": "Develop a project plan for a small infrastructure project",            "duration": "5 weeks"},
"Quantity Surveying":                     {"course": "Quantity Surveying - Udemy",                                   "platform": "https://www.udemy.com",                          "project": "Prepare full quantity take-offs for a 2BHK house",                     "duration": "3 weeks"},

# ── CHEMICAL / AEROSPACE — fill in the gaps ──────────────────
"Process Simulation":                     {"course": "Process Simulation with Aspen - Udemy",                        "platform": "https://www.udemy.com",                          "project": "Simulate a distillation column in Aspen HYSYS",                        "duration": "4 weeks"},
"HYSYS":                                  {"course": "Aspen HYSYS Beginner Course - Udemy",                          "platform": "https://www.udemy.com",                          "project": "Model a complete process flow diagram in HYSYS",                       "duration": "4 weeks"},
"Chemical Process Design":                {"course": "Chemical Process Design - NPTEL",                              "platform": "https://nptel.ac.in",                            "project": "Design a continuous reactor for a chosen reaction",                    "duration": "5 weeks"},
"Mass Transfer":                          {"course": "Mass Transfer - NPTEL",                                        "platform": "https://nptel.ac.in",                            "project": "Solve 15 absorption and distillation problems",                        "duration": "5 weeks"},
"Process Optimization":                   {"course": "Process Optimization - Coursera",                              "platform": "https://www.coursera.org",                       "project": "Optimise a small chemical plant for cost & yield",                     "duration": "4 weeks"},
"Aerodynamics":                           {"course": "Introduction to Aerodynamics - MIT OpenCourseware",            "platform": "https://ocw.mit.edu",                            "project": "Analyse lift and drag of 3 airfoil shapes",                            "duration": "5 weeks"},
"Propulsion Systems":                     {"course": "Aerospace Propulsion - NPTEL",                                 "platform": "https://nptel.ac.in",                            "project": "Compare 3 types of jet engines with performance charts",               "duration": "5 weeks"},
"Avionics":                               {"course": "Avionics Systems - NPTEL",                                     "platform": "https://nptel.ac.in",                            "project": "Document 5 modern avionics systems with block diagrams",               "duration": "4 weeks"},
"Flight Mechanics":                       {"course": "Flight Mechanics - NPTEL",                                     "platform": "https://nptel.ac.in",                            "project": "Solve 20 aircraft stability and performance problems",                 "duration": "5 weeks"},
"Safety Management":                      {"course": "Workplace Safety - Coursera",                                  "platform": "https://www.coursera.org",                       "project": "Develop a safety SOP for a small chemical plant",                      "duration": "3 weeks"},
"Process Optimization":                   {"course": "Process Optimization - Coursera",                              "platform": "https://www.coursera.org",                       "project": "Identify and reduce waste in a sample process flow",                   "duration": "4 weeks"},

# ── MISC SHARED SKILLS ───────────────────────────────────────
"Cloud Computing":                        {"course": "AWS Cloud Practitioner - Udemy",                               "platform": "https://www.udemy.com",                          "project": "Deploy a Flask app on AWS EC2",                                        "duration": "5 weeks"},
"Cybersecurity":                          {"course": "Introduction to Cybersecurity - Coursera (NYU)",               "platform": "https://www.coursera.org",                       "project": "Audit a small network for common vulnerabilities",                     "duration": "5 weeks"},
"Networking":                             {"course": "Computer Networking - Coursera (Google)",                      "platform": "https://www.coursera.org",                       "project": "Configure a small office LAN with router and switch",                  "duration": "4 weeks"},
"Windows Server":                         {"course": "Windows Server Administration - Udemy",                        "platform": "https://www.udemy.com",                          "project": "Set up Active Directory in a virtual environment",                     "duration": "4 weeks"},
"Hospitality Marketing":                  {"course": "Hospitality Marketing - Coursera",                             "platform": "https://www.coursera.org",                       "project": "Plan a marketing campaign for a small hotel",                          "duration": "3 weeks"},
}

# ============================================================
#  DATABASE MODELS
# ============================================================
class Student(UserMixin, db.Model):
    __tablename__ = 'students'
    id          = db.Column(db.Integer, primary_key=True)
    name        = db.Column(db.String(100), nullable=False)
    email       = db.Column(db.String(100), unique=True, nullable=False)
    password    = db.Column(db.String(200), nullable=False)
    degree      = db.Column(db.String(50),  nullable=False)
    stream      = db.Column(db.String(50),  nullable=True)
    created_at  = db.Column(db.DateTime,    default=datetime.utcnow)
    predictions = db.relationship('Prediction', backref='student', lazy=True)
    def get_id(self): return f's_{self.id}'

class Recruiter(UserMixin, db.Model):
    __tablename__ = 'recruiters'
    id           = db.Column(db.Integer, primary_key=True)
    name         = db.Column(db.String(100), nullable=False)
    email        = db.Column(db.String(100), unique=True, nullable=False)
    password     = db.Column(db.String(200), nullable=False)
    company_name = db.Column(db.String(150), nullable=False)
    industry     = db.Column(db.String(100), nullable=False)
    created_at   = db.Column(db.DateTime,    default=datetime.utcnow)
    shortlists   = db.relationship('Shortlist', backref='recruiter', lazy=True)
    def get_id(self): return f'r_{self.id}'

class Admin(UserMixin, db.Model):
    __tablename__ = 'admins'
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(100), nullable=False)
    email      = db.Column(db.String(100), unique=True, nullable=False)
    password   = db.Column(db.String(200), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    def get_id(self): return f'a_{self.id}'

class Prediction(db.Model):
    __tablename__ = 'predictions'
    id                  = db.Column(db.Integer, primary_key=True)
    student_id          = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    desired_role        = db.Column(db.String(100), nullable=False)
    cgpa                = db.Column(db.Float,   nullable=False)
    skills              = db.Column(db.String(500), nullable=False)
    internships         = db.Column(db.Integer, default=0)
    projects            = db.Column(db.Integer, default=0)
    backlogs            = db.Column(db.Integer, default=0)
    communication_score = db.Column(db.Integer, default=5)
    placement_chance    = db.Column(db.Float,   nullable=False)
    skill_match         = db.Column(db.Float,   nullable=False)
    missing_skills      = db.Column(db.String(500), nullable=False)
    created_at          = db.Column(db.DateTime, default=datetime.utcnow)

class Shortlist(db.Model):
    __tablename__ = 'shortlists'
    id           = db.Column(db.Integer, primary_key=True)
    recruiter_id = db.Column(db.Integer, db.ForeignKey('recruiters.id'), nullable=False)
    student_id   = db.Column(db.Integer, db.ForeignKey('students.id'),   nullable=False)
    note         = db.Column(db.String(300), nullable=True)
    status       = db.Column(db.String(20),  default='Pending')
    joining_date  = db.Column(db.Date, nullable=True)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)
    student      = db.relationship('Student', backref='shortlists')
class Degree(db.Model):
    __tablename__ = 'degrees'
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(80),  unique=True, nullable=False)
    full_name  = db.Column(db.String(200), nullable=False)
    level      = db.Column(db.String(50),  nullable=False, default='UG')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    job_roles  = db.relationship('DegreeJobRole', backref='degree',
                                  lazy=True, cascade='all, delete-orphan')
    skills     = db.relationship('DegreeSkill',   backref='degree',
                                  lazy=True, cascade='all, delete-orphan')

class DegreeJobRole(db.Model):
    __tablename__ = 'degree_job_roles'
    id        = db.Column(db.Integer, primary_key=True)
    degree_id = db.Column(db.Integer, db.ForeignKey('degrees.id'), nullable=False)
    name      = db.Column(db.String(200), nullable=False)

class DegreeSkill(db.Model):
    __tablename__ = 'degree_skills'
    id        = db.Column(db.Integer, primary_key=True)
    degree_id = db.Column(db.Integer, db.ForeignKey('degrees.id'), nullable=False)
    name      = db.Column(db.String(200), nullable=False)

def reload_degree_data():
    """Refresh in-memory DEGREE_DATA + ALL_DEGREES + COMPETITIVE_EXAM_ROLES from DB."""
    global DEGREE_DATA, ALL_DEGREES, COMPETITIVE_EXAM_ROLES
    DEGREE_DATA = {}
    for d in Degree.query.order_by(Degree.name).all():
        DEGREE_DATA[d.name] = {
            'full_name': d.full_name,
            'level': d.level,
            'job_roles': [r.name for r in d.job_roles],
            'recommended_skills': [s.name for s in d.skills],
        }
    ALL_DEGREES = sorted(DEGREE_DATA.keys())
    COMPETITIVE_EXAM_ROLES = []
    for _, _data in DEGREE_DATA.items():
        if _data.get('level') == 'Competitive Exam':
            COMPETITIVE_EXAM_ROLES.extend(_data['job_roles'])

def seed_degrees_if_empty():
    """Populate the degrees table from DEGREE_DATA_SEED.
    On the first run it seeds everything; on subsequent runs it only
    inserts NEW degrees that are not already in the database, so that
    adding more entries to DEGREE_DATA_SEED simply propagates on restart."""
    existing_names = {d.name for d in Degree.query.all()}
    added = 0
    for name, data in DEGREE_DATA_SEED.items():
        if name in existing_names:
            continue
        d = Degree(name=name,
                   full_name=data['full_name'],
                   level=data.get('level', 'UG'))
        db.session.add(d)
        db.session.flush()
        for role in data.get('job_roles', []):
            db.session.add(DegreeJobRole(degree_id=d.id, name=role))
        for skill in data.get('recommended_skills', []):
            db.session.add(DegreeSkill(degree_id=d.id, name=skill))
        added += 1
    if added > 0:
        db.session.commit()
        print(f'🌱 Seeded {added} new degree(s) into the database')


@login_manager.user_loader
def load_user(uid):
    if uid.startswith('s_'): return Student.query.get(int(uid[2:]))
    if uid.startswith('r_'): return Recruiter.query.get(int(uid[2:]))
    if uid.startswith('a_'): return Admin.query.get(int(uid[2:]))
    return None

def student_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if not isinstance(current_user, Student):
            flash('Student access only.', 'error')
            return redirect(url_for('login'))
        return f(*a, **kw)
    return dec

def recruiter_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if not isinstance(current_user, Recruiter):
            flash('Recruiter access only.', 'error')
            return redirect(url_for('login'))
        return f(*a, **kw)
    return dec

def admin_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if not isinstance(current_user, Admin):
            flash('Admin access only.', 'error')
            return redirect(url_for('login'))
        return f(*a, **kw)
    return dec

# ============================================================
#  LOAD CSV & TRAIN MODEL
# ============================================================
df     = pd.read_csv('student_data.csv')
job_db = pd.read_csv('job_skills_database.csv')

le_stream = LabelEncoder()
le_job    = LabelEncoder()
le_placed = LabelEncoder()

df['stream_encoded']   = le_stream.fit_transform(df['stream'])
df['job_role_encoded'] = le_job.fit_transform(df['desired_job_role'])
df['placed_encoded']   = le_placed.fit_transform(df['placed'])
df['skill_count']      = df['skills'].apply(lambda x: len(str(x).split(';')))

FEATURES = ['cgpa', 'stream_encoded', 'job_role_encoded',
            'internships', 'projects', 'backlogs',
            'communication_score', 'skill_count']

X = df[FEATURES]
y = df['placed_encoded']
X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.2, random_state=42)
rf_model = RandomForestClassifier(
    n_estimators=500, max_depth=10, random_state=42, class_weight='balanced')
rf_model.fit(X_train, y_train)
print(f'✅ Model ready | Accuracy: {accuracy_score(y_test, rf_model.predict(X_test))*100:.2f}%')

ALL_JOB_ROLES = sorted(job_db['job_role'].tolist())

# ============================================================
#  HELPERS
# ============================================================
def analyze_skill_gap(student_skills, desired_role, fallback_skills=None):
    row = job_db[job_db['job_role'] == desired_role]
    if row.empty:
        # Role not in CSV — use degree's recommended skills as fallback
        required  = fallback_skills if fallback_skills else ['Communication', 'MS Office', 'Critical Thinking']
        good_have = []
        min_cgpa  = 6.0
    else:
        required  = [s.strip() for s in row['required_skills'].values[0].split(';')]
        good_have = [s.strip() for s in row['good_to_have_skills'].values[0].split(';')]
        min_cgpa  = float(row['min_cgpa'].values[0])
    ignore    = {'nothing', 'none', 'nil', 'na', 'n/a', 'no skills', '-', 'null', ''}
    s_skills  = {s.strip().lower() for s in student_skills.split(';')
                 if s.strip().lower() not in ignore}
    matched = [s for s in required  if s.lower() in s_skills]
    missing = [s for s in required  if s.lower() not in s_skills]
    miss_g  = [s for s in good_have if s.lower() not in s_skills]
    pct     = round(len(matched) / len(required) * 100, 2) if required else 0
    return {'required': required, 'matched': matched,
            'missing': missing, 'missing_good': miss_g,
            'match_pct': pct, 'min_cgpa': min_cgpa}

def get_recs(missing_skills):
    recs = []
    for skill in missing_skills:
        r = RECOMMENDATIONS.get(skill, {
            'course':   f'Search "{skill}" on YouTube',
            'platform': 'https://www.youtube.com',
            'project':  f'Build a small project using {skill}',
            'duration': '2-3 weeks'
        })
        recs.append({'skill': skill, **r})
    return recs

def run_prediction(student, cgpa, skills, internships,
                   projects, backlogs, comm, role):
    stream = student.stream if student.stream else student.degree
    s_enc  = le_stream.transform([stream])[0] if stream in le_stream.classes_ else 0
    j_enc  = le_job.transform([role])[0] if role in le_job.classes_ else 0
    ignore = {'nothing', 'none', 'nil', 'na', 'n/a', 'no skills', '-', 'null', ''}
    sc     = len([s for s in skills.split(';') if s.strip().lower() not in ignore])
    X_in   = pd.DataFrame([[cgpa, s_enc, j_enc, internships, projects,
                             backlogs, comm, sc]], columns=FEATURES)
    raw = rf_model.predict_proba(X_in)[0][1] * 100
    if   sc == 0 and internships == 0 and projects == 0: prob = round(raw * 0.20, 2)
    elif sc == 0 and internships == 0:                   prob = round(raw * 0.35, 2)
    elif sc == 0:                                        prob = round(raw * 0.50, 2)
    else:                                                prob = round(raw, 2)
    return prob

# ── Reset password token helpers ──
def get_reset_serializer():
    return URLSafeTimedSerializer(app.config['SECRET_KEY'])

def generate_reset_token(email):
    """Generate a time-limited token for password reset."""
    return get_reset_serializer().dumps(email, salt='password-reset')

def verify_reset_token(token, max_age_seconds=3600):
    """Verify the token and return the email; None if invalid/expired."""
    try:
        return get_reset_serializer().loads(token, salt='password-reset',
                                            max_age=max_age_seconds)
    except Exception:
        return None


def gemini_roadmap(name, role, missing, chance):
    if not gemini_client:
        return '⚠ AI roadmap unavailable — Gemini client not initialized. Check your API key.'
    try:
        prompt = f"""You are a friendly career counselor. Generate a personalized 3-month career roadmap for a student.

Student Name: {name}
Target Role: {role}
Current Placement Chance: {chance}%
Skills Still Needed: {', '.join(missing) if missing else 'None — already has all required skills!'}

Please provide:
- Month 1 Plan: Focus on foundation skills (2-3 specific topics with what to do)
- Month 2 Plan: Build intermediate skills + 1-2 mini projects
- Month 3 Plan: Advanced topics + portfolio building + interview prep
- Top 3 Companies or Organizations to target for this role
- One motivational tip to keep going

Be concise, practical, and encouraging. Use plain text formatting, no markdown asterisks. Keep it under 400 words."""

        response = gemini_client.models.generate_content(
            model=GEMINI_MODEL_NAME,
            contents=prompt
        )
        return response.text
    except Exception as e:
        # Print actual error to terminal for debugging
        print(f'⚠ Gemini API call failed for {role}: {e}')
        return f'⚠ AI roadmap temporarily unavailable.\n\nError: {str(e)[:200]}\n\nFor now, focus on the skill recommendations above and work through them month by month!'

# ============================================================
#  API
# ============================================================
@app.route('/api/degree_info/<path:degree>')
def api_degree_info(degree):
    data = DEGREE_DATA.get(degree)
    if not data:
        return jsonify({'error': 'Not found'}), 404
    return jsonify(data)

# ============================================================
#  PUBLIC ROUTES
# ============================================================
@app.route('/')
def home():
    return render_template('home.html',
        total_students=Student.query.count(),
        total_recruiters=Recruiter.query.count(),
        total_predictions=Prediction.query.count())

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        if isinstance(current_user, Admin):     return redirect(url_for('admin_dashboard'))
        if isinstance(current_user, Recruiter): return redirect(url_for('recruiter_dashboard'))
        return redirect(url_for('student_dashboard'))
    if request.method == 'POST':
        role  = request.form.get('role', 'student')
        email = request.form.get('email', '').strip()
        pw    = request.form.get('password', '')
        if role == 'student':
            u = Student.query.filter_by(email=email).first()
            if u and check_password_hash(u.password, pw):
                login_user(u); return redirect(url_for('student_dashboard'))
            flash('Invalid student email or password.', 'error')
            return render_template('login.html', active_tab='student')
        elif role == 'recruiter':
            u = Recruiter.query.filter_by(email=email).first()
            if u and check_password_hash(u.password, pw):
                login_user(u); return redirect(url_for('recruiter_dashboard'))
            flash('Invalid recruiter email or password.', 'error')
            return render_template('login.html', active_tab='recruiter')
        elif role == 'admin':
            # Step 1: Verify the secret key first
            submitted_key = request.form.get('admin_key', '').strip()
            if submitted_key != ADMIN_SECRET_KEY:
                flash('Invalid admin secret key. Access denied.', 'error')
                return render_template('login.html', active_tab='admin')
            # Step 2: Verify email + password
            u = Admin.query.filter_by(email=email).first()
            if u and check_password_hash(u.password, pw):
                login_user(u); return redirect(url_for('admin_dashboard'))
            flash('Invalid admin credentials.', 'error')
            return render_template('login.html', active_tab='admin')
    return render_template('login.html',
                           active_tab=request.args.get('tab', 'student'))

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out successfully.', 'info')
    return redirect(url_for('home'))

# ============================================================
#  FORGOT PASSWORD ROUTES
# ============================================================
@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        # Look up the user in students or recruiters
        student = Student.query.filter_by(email=email).first()
        recruiter = Recruiter.query.filter_by(email=email).first()
        user = student or recruiter

        if user:
            token = generate_reset_token(email)
            reset_url = url_for('reset_password', token=token, _external=True)
            body = (
                f"Hi {user.name},\n\n"
                f"You requested to reset your password for Placement Predictor.\n\n"
                f"Click the link below to reset your password (valid for 1 hour):\n"
                f"{reset_url}\n\n"
                f"If you didn't request this, please ignore this email — your password will remain unchanged.\n\n"
                f"Best wishes,\n"
                f"Placement Predictor Team"
            )
            try:
                msg = Message(
                    subject='🔐 Reset your Placement Predictor password',
                    recipients=[email],
                    body=body
                )
                mail.send(msg)
                flash(f'Password reset link sent to {email}! Please check your inbox (and spam folder).', 'success')
            except Exception as e:
                print(f'⚠ Email send failed: {e}')
                flash('Could not send the reset email right now. Please try again later or contact the admin.', 'error')
        else:
            # Don't reveal whether the email exists — show same message either way
            flash(f'If an account exists for {email}, a reset link has been sent. Please check your inbox.', 'info')

        return redirect(url_for('login'))

    return render_template('forgot_password.html')


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    email = verify_reset_token(token)
    if not email:
        flash('This reset link is invalid or has expired. Please request a new one.', 'error')
        return redirect(url_for('forgot_password'))

    # Find the user
    student = Student.query.filter_by(email=email).first()
    recruiter = Recruiter.query.filter_by(email=email).first()
    user = student or recruiter
    if not user:
        flash('User account not found.', 'error')
        return redirect(url_for('login'))

    if request.method == 'POST':
        new_password = request.form.get('password', '')
        confirm = request.form.get('confirm', '')
        if new_password != confirm:
            flash('Passwords do not match. Please try again.', 'error')
            return render_template('reset_password.html', token=token, email=email)
        if len(new_password) < 6:
            flash('Password must be at least 6 characters long.', 'error')
            return render_template('reset_password.html', token=token, email=email)

        user.password = generate_password_hash(new_password)
        db.session.commit()
        flash('🎉 Password reset successful! Please login with your new password.', 'success')
        return redirect(url_for('login'))

    return render_template('reset_password.html', token=token, email=email)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name   = request.form['name'].strip()
        email  = request.form['email'].strip()
        pw     = request.form['password']
        degree = request.form['degree']
        stream = request.form.get('stream', '').strip()
        if Student.query.filter_by(email=email).first():
            flash('Email already registered!', 'error')
            return redirect(url_for('register'))
        db.session.add(Student(name=name, email=email,
            password=generate_password_hash(pw),
            degree=degree, stream=stream))
        db.session.commit()
        flash('Account created! Please login.', 'success')
        return redirect(url_for('login'))
    return render_template('register.html',
        degrees=ALL_DEGREES, degree_data=DEGREE_DATA)

@app.route('/recruiter/register', methods=['GET', 'POST'])
def recruiter_register():
    if request.method == 'POST':
        name    = request.form['name'].strip()
        email   = request.form['email'].strip()
        pw      = request.form['password']
        company = request.form['company_name'].strip()
        industry= request.form['industry']
        if Recruiter.query.filter_by(email=email).first():
            flash('Email already registered!', 'error')
            return redirect(url_for('recruiter_register'))
        db.session.add(Recruiter(name=name, email=email,
            password=generate_password_hash(pw),
            company_name=company, industry=industry))
        db.session.commit()
        flash('Recruiter account created! Please login.', 'success')
        return redirect(url_for('login') + '?tab=recruiter')
    return render_template('recruiter_register.html')

# ============================================================
#  STUDENT ROUTES
# ============================================================
@app.route('/student/dashboard')
@login_required
@student_required
def student_dashboard():
    deg_info  = DEGREE_DATA.get(current_user.degree, {})
    preds     = Prediction.query.filter_by(student_id=current_user.id)\
                                .order_by(Prediction.created_at.desc()).all()
    # Check if this student has been hired by any recruiter
    hired_sl = Shortlist.query.filter_by(
        student_id=current_user.id, status='Hired').first()
    hired_info = None
    if hired_sl:
        hired_info = {
            'company': hired_sl.recruiter.company_name,
            'joining_date': hired_sl.joining_date,
            'recruiter_name': hired_sl.recruiter.name,
        }
    return render_template('student_dashboard.html',
        user=current_user, predictions=preds,
        job_roles=deg_info.get('job_roles', ALL_JOB_ROLES),
        competitive_roles=COMPETITIVE_EXAM_ROLES,
        rec_skills=deg_info.get('recommended_skills', []),
        degree_info=deg_info,
        hired_info=hired_info)

@app.route('/student/predict', methods=['POST'])
@login_required
@student_required
def student_predict():
    cgpa   = float(request.form['cgpa'])
    skills = request.form['skills']
    interns= int(request.form['internships'])
    projs  = int(request.form['projects'])
    backs  = int(request.form['backlogs'])
    comm   = int(request.form['communication_score'])

    # Get ALL selected roles (multi-select via name="desired_roles")
    selected_roles = request.form.getlist('desired_roles')
    # Clean: remove blanks, deduplicate, keep order
    selected_roles = list(dict.fromkeys([r.strip() for r in selected_roles if r.strip()]))

    if not selected_roles:
        flash('Please select at least one role.', 'error')
        return redirect(url_for('student_dashboard'))

    deg_skills = DEGREE_DATA.get(current_user.degree, {}).get('recommended_skills', [])

    # Run prediction for each selected role
    role_results = []
    for role in selected_roles:
        prob = run_prediction(current_user, cgpa, skills,
                              interns, projs, backs, comm, role)
        status_text = 'Likely to be Placed' if prob >= 50 else 'Needs Improvement'
        role_skills = get_skills_for_role(role, deg_skills)
        gap     = analyze_skill_gap(skills, role, fallback_skills=role_skills)
        recs    = get_recs(gap['missing'])
        roadmap = gemini_roadmap(current_user.name, role, gap['missing'], prob)

        # Save each role's prediction as a separate row in the DB
        db.session.add(Prediction(
            student_id=current_user.id, desired_role=role,
            cgpa=cgpa, skills=skills, internships=interns,
            projects=projs, backlogs=backs, communication_score=comm,
            placement_chance=prob, skill_match=gap['match_pct'],
            missing_skills=', '.join(gap['missing'])))

        role_results.append({
            'role': role,
            'prob': prob,
            'status': status_text,
            'gap': gap,
            'recs': recs,
            'roadmap': roadmap,
            'cgpa_warn': cgpa < gap['min_cgpa'],
        })

    db.session.commit()

    # Sort by placement chance descending — best matches on top
    role_results.sort(key=lambda x: x['prob'], reverse=True)

    deg_info = DEGREE_DATA.get(current_user.degree, {})
    preds    = Prediction.query.filter_by(student_id=current_user.id)\
                               .order_by(Prediction.created_at.desc()).all()

    hired_sl = Shortlist.query.filter_by(
        student_id=current_user.id, status='Hired').first()
    hired_info = None
    if hired_sl:
        hired_info = {
            'company': hired_sl.recruiter.company_name,
            'joining_date': hired_sl.joining_date,
            'recruiter_name': hired_sl.recruiter.name,
        }

    return render_template('student_dashboard.html',
        user=current_user, predictions=preds,
        job_roles=deg_info.get('job_roles', ALL_JOB_ROLES),
        competitive_roles=COMPETITIVE_EXAM_ROLES,
        rec_skills=deg_info.get('recommended_skills', []),
        degree_info=deg_info,
        hired_info=hired_info,
        result=True,
        role_results=role_results,
        skills=skills, cgpa=cgpa)

@app.route('/student/download', methods=['POST'])
@login_required
@student_required
def student_download():
    f      = request.form
    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=letter)
    styles = getSampleStyleSheet()
    story  = []

    t_s = ParagraphStyle('T', parent=styles['Title'], fontSize=22,
                          textColor=colors.HexColor('#667eea'), alignment=1, spaceAfter=20)
    h_s = ParagraphStyle('H', parent=styles['Heading2'], fontSize=13,
                          textColor=colors.HexColor('#764ba2'), spaceAfter=8, spaceBefore=10)
    c_s = ParagraphStyle('C', parent=styles['Normal'], fontSize=11,
                          textColor=colors.HexColor('#667eea'), alignment=1)

    story += [
        Paragraph('Career Placement Report', t_s),
        Paragraph(f"Student Name : {f['name']}", styles['Normal']),
        Paragraph(f"Desired Role : {f['desired_role']}", styles['Normal']),
        Paragraph(f"Report Date  : {datetime.now().strftime('%d %B %Y, %I:%M %p')}", styles['Normal']),

        Spacer(1, 14),
        Paragraph('Placement Prediction', h_s),
        Paragraph(f"Placement Chance : {f['probability']}%", styles['Normal']),
        Paragraph(f"Status : {f['status']}", styles['Normal']),
        Spacer(1, 12),
        Paragraph('Skills You Have', h_s),
        Paragraph(f['matched_skills'] or 'None', styles['Normal']),
        Spacer(1, 12),
        Paragraph('Skills You Require', h_s),
        Paragraph(f['missing_skills'] or 'None', styles['Normal']),
        Spacer(1, 12),
        Paragraph('Recommendations', h_s),
    ]

    for line in f['recommendations_text'].split('|'):
        if line.strip():
            story.append(Paragraph(f'• {line.strip()}', styles['Normal']))

    # ── AI Career Roadmap ─────────────────────────────────────
    roadmap_text = f.get('llm_roadmap', '').strip()
    if roadmap_text:
        # Clean up any markdown that ReportLab can't render
        roadmap_text = roadmap_text.replace('**', '').replace('*', '')
        story += [
            Spacer(1, 14),
            Paragraph('AI Career Roadmap', h_s),
        ]
        # Render each line as its own paragraph so formatting looks neat
        for line in roadmap_text.split('\n'):
            line = line.strip()
            if line:
                # Escape XML special characters so ReportLab parses safely
                safe_line = (line.replace('&', '&amp;')
                                  .replace('<', '&lt;')
                                  .replace('>', '&gt;'))
                story.append(Paragraph(safe_line, styles['Normal']))
            else:
                story.append(Spacer(1, 6))

    story += [
        Spacer(1, 24),
        Paragraph('Best Wishes for Your Future!', c_s),
        Paragraph('Keep learning, keep growing, and never stop believing in yourself!', c_s),
        Spacer(1, 14),
        Paragraph('Generated by Student Career Placement Predictor', styles['Italic'])
    ]

    doc.build(story)
    buf.seek(0)
    safe_name = f['name'].replace(' ', '_')
    date_str  = datetime.now().strftime('%Y-%m-%d')
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f"attachment; filename={safe_name}_placement_report_{date_str}.pdf"
    return resp

@app.route('/student/prediction/<int:pid>/delete', methods=['POST'])
@login_required
@student_required
def student_prediction_delete(pid):
    # Security: only allow students to delete their OWN predictions
    pred = Prediction.query.filter_by(id=pid, student_id=current_user.id).first()
    if not pred:
        flash('Prediction not found or not yours to delete.', 'error')
        return redirect(url_for('student_dashboard'))
    role = pred.desired_role
    db.session.delete(pred)
    db.session.commit()
    flash(f'Prediction for "{role}" deleted successfully.', 'success')
    return redirect(url_for('student_dashboard'))

@app.route('/student/predictions/delete', methods=['POST'])
@login_required
@student_required
def student_predictions_bulk_delete():
    # Security: only delete predictions that belong to the current student
    ids = request.form.getlist('prediction_ids')
    if not ids:
        flash('No predictions selected.', 'info')
        return redirect(url_for('student_dashboard'))
    deleted = Prediction.query.filter(
        Prediction.id.in_(ids),
        Prediction.student_id == current_user.id
    ).delete(synchronize_session=False)
    db.session.commit()
    flash(f'{deleted} prediction(s) deleted successfully.', 'success')
    return redirect(url_for('student_dashboard'))
# ============================================================
#  RECRUITER ROUTES
# ============================================================
@app.route('/recruiter/dashboard')
@login_required
@recruiter_required
def recruiter_dashboard():
    skill_f  = request.args.get('skill',    '').strip()
    role_f   = request.args.get('role',     '').strip()
    degree_f = request.args.get('degree',   '').strip()
    min_cgpa = request.args.get('min_cgpa', type=float, default=0.0)
    sort_by  = request.args.get('sort',     'placement_desc')

    subq = db.session.query(
        Prediction.student_id,
        db.func.max(Prediction.id).label('max_id')
    ).group_by(Prediction.student_id).subquery()

    q = (db.session.query(Student, Prediction)
         .join(Prediction, Prediction.student_id == Student.id)
         .join(subq, (subq.c.student_id == Prediction.student_id) &
                     (subq.c.max_id == Prediction.id))
         .filter(Prediction.cgpa >= min_cgpa))

    if role_f:   q = q.filter(Prediction.desired_role == role_f)
    if degree_f: q = q.filter(Student.degree == degree_f)
    if skill_f:  q = q.filter(Prediction.skills.ilike(f'%{skill_f}%'))

    if sort_by == 'cgpa_desc':  q = q.order_by(Prediction.cgpa.desc())
    elif sort_by == 'name_asc': q = q.order_by(Student.name.asc())
    else:                       q = q.order_by(Prediction.placement_chance.desc())

    students        = q.all()
    shortlisted_ids = {s.student_id for s in current_user.shortlists}
    all_roles       = sorted({p.desired_role for _, p in
                              db.session.query(Student, Prediction)
                              .join(Prediction, Prediction.student_id == Student.id).all()})
    return render_template('recruiter_dashboard.html',
        recruiter=current_user, students=students,
        shortlisted_ids=shortlisted_ids, all_roles=all_roles,
        all_degrees=ALL_DEGREES,
        filters={'skill': skill_f, 'role': role_f,
                 'degree': degree_f, 'min_cgpa': min_cgpa, 'sort': sort_by})

@app.route('/recruiter/shortlist/<int:sid>', methods=['POST'])
@login_required
@recruiter_required
def toggle_shortlist(sid):
    ex = Shortlist.query.filter_by(
        recruiter_id=current_user.id, student_id=sid).first()
    if ex:
        db.session.delete(ex)
        db.session.commit()
        return jsonify({'status': 'removed'})
    db.session.add(Shortlist(recruiter_id=current_user.id,
                             student_id=sid,
                             note=request.form.get('note', '')))
    db.session.commit()
    return jsonify({'status': 'added'})

@app.route('/recruiter/shortlist/<int:sid>/update', methods=['POST'])
@login_required
@recruiter_required
def update_shortlist(sid):
    sl = Shortlist.query.filter_by(
        recruiter_id=current_user.id, student_id=sid).first()
    if not sl:
        return jsonify({'error': 'Not in your shortlist'}), 404
    if 'note' in request.form:
        sl.note = request.form['note'].strip()[:300]
    if 'status' in request.form:
        valid = ['Pending', 'Contacted', 'Interviewed', 'Hired', 'Rejected']
        if request.form['status'] in valid:
            sl.status = request.form['status']
    if 'joining_date' in request.form:
        date_str = request.form['joining_date'].strip()
        if date_str:
            try:
                sl.joining_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        else:
            sl.joining_date = None
    # Auto-set joining date to today if status just became Hired
    if sl.status == 'Hired' and not sl.joining_date:
        sl.joining_date = datetime.utcnow().date()
    db.session.commit()
    return jsonify({
        'ok': True,
        'note': sl.note or '',
        'shortlist_status': sl.status,
        'joining_date': sl.joining_date.strftime('%Y-%m-%d') if sl.joining_date else ''
    })

@app.route('/recruiter/shortlists')
@login_required
@recruiter_required
def recruiter_shortlists():
    sl = Shortlist.query.filter_by(recruiter_id=current_user.id)\
                        .order_by(Shortlist.created_at.desc()).all()
    return render_template('recruiter_shortlists.html',
        recruiter=current_user, shortlists=sl)

@app.route('/recruiter/student/<int:sid>')
@login_required
@recruiter_required
def recruiter_view_student(sid):
    student = Student.query.get_or_404(sid)
    preds   = Prediction.query.filter_by(student_id=sid)\
                              .order_by(Prediction.created_at.desc()).all()
    is_sl   = Shortlist.query.filter_by(
        recruiter_id=current_user.id, student_id=sid).first() is not None
    return render_template('recruiter_student_profile.html',
        recruiter=current_user, student=student,
        predictions=preds, is_shortlisted=is_sl,
        degree_data=DEGREE_DATA)

# ============================================================
#  ADMIN ROUTES
# ============================================================
@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    stats = {
        'total_students':    Student.query.count(),
        'total_recruiters':  Recruiter.query.count(),
        'total_predictions': Prediction.query.count(),
        'total_shortlists':  Shortlist.query.count(),
        'avg_placement': round(db.session.query(
            db.func.avg(Prediction.placement_chance)).scalar() or 0, 2),
    }

    # Chart data — degree distribution
    degree_counts = db.session.query(
        Student.degree, db.func.count(Student.id).label('count')
    ).group_by(Student.degree).order_by(db.desc('count')).all()

    # Chart data — avg placement chance per degree
    avg_by_degree = db.session.query(
        Student.degree, db.func.avg(Prediction.placement_chance).label('avg_c')
    ).join(Prediction, Prediction.student_id == Student.id)\
     .group_by(Student.degree).order_by(db.desc('avg_c')).all()

    chart_data = {
        'degree_dist_labels': [d for d, _ in degree_counts],
        'degree_dist_values': [int(c) for _, c in degree_counts],
        'avg_placement_labels': [d for d, _ in avg_by_degree],
        'avg_placement_values': [round(float(v), 2) for _, v in avg_by_degree],
    }

    return render_template('admin_dashboard.html',
        admin=current_user,
        students=Student.query.order_by(Student.created_at.desc()).all(),
        recruiters=Recruiter.query.order_by(Recruiter.created_at.desc()).all(),
        predictions=Prediction.query.order_by(
            Prediction.created_at.desc()).limit(20).all(),
        stats=stats,
        chart_data=chart_data)

@app.route('/admin/delete_student/<int:sid>', methods=['POST'])
@login_required
@admin_required
def admin_delete_student(sid):
    s = Student.query.get_or_404(sid)
    Prediction.query.filter_by(student_id=sid).delete()
    Shortlist.query.filter_by(student_id=sid).delete()
    db.session.delete(s)
    db.session.commit()
    flash(f'Student {s.name} deleted.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete_recruiter/<int:rid>', methods=['POST'])
@login_required
@admin_required
def admin_delete_recruiter(rid):
    r = Recruiter.query.get_or_404(rid)
    Shortlist.query.filter_by(recruiter_id=rid).delete()
    db.session.delete(r)
    db.session.commit()
    flash(f'Recruiter {r.name} deleted.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/export/<dtype>')
@login_required
@admin_required
def admin_export(dtype):
    wb = Workbook()
    ws = wb.active
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')

    if dtype == 'students':
        ws.title = 'Students'
        headers = ['ID', 'Name', 'Email', 'Degree', 'Specialisation', 'Joined',
                   'Hired Company', 'Joining Date', 'Hired By Recruiter']
        ws.append(headers)
        for s in Student.query.order_by(Student.id).all():
            hired = Shortlist.query.filter_by(student_id=s.id, status='Hired').first()
            company = hired.recruiter.company_name if hired else ''
            join_date = hired.joining_date.strftime('%d %b %Y') if hired and hired.joining_date else ''
            recruiter_name = hired.recruiter.name if hired else ''
            ws.append([s.id, s.name, s.email, s.degree, s.stream or '—',
                       s.created_at.strftime('%d %b %Y %I:%M %p'),
                       company, join_date, recruiter_name])
        filename = 'students_export.xlsx'

    elif dtype == 'recruiters':
        ws.title = 'Recruiters'
        headers = ['ID', 'Name', 'Email', 'Company', 'Industry', 'Joined']
        ws.append(headers)
        for r in Recruiter.query.order_by(Recruiter.id).all():
            ws.append([r.id, r.name, r.email, r.company_name, r.industry,
                       r.created_at.strftime('%d %b %Y %I:%M %p')])
        filename = 'recruiters_export.xlsx'

    elif dtype == 'predictions':
        ws.title = 'Predictions'
        headers = ['ID', 'Student', 'Email', 'Degree', 'Role', 'CGPA', 'Internships',
                   'Projects', 'Backlogs', 'Comm Score', 'Skills',
                   'Placement Chance %', 'Skill Match %', 'Missing Skills', 'Date']
        ws.append(headers)
        for p in Prediction.query.order_by(Prediction.id).all():
            ws.append([p.id, p.student.name, p.student.email, p.student.degree,
                       p.desired_role, p.cgpa, p.internships, p.projects, p.backlogs,
                       p.communication_score, p.skills,
                       p.placement_chance, p.skill_match, p.missing_skills,
                       p.created_at.strftime('%d %b %Y %I:%M %p')])
        filename = 'predictions_export.xlsx'

    else:
        flash('Invalid export type', 'error')
        return redirect(url_for('admin_dashboard'))

    # Style header row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Auto-size columns
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return resp

@app.route('/admin/create', methods=['GET', 'POST'])
def admin_create():
    if Admin.query.count() > 0:
        flash('Admin already exists. Please login.', 'info')
        return redirect(url_for('login') + '?tab=admin')
    if request.method == 'POST':
        db.session.add(Admin(
            name=request.form['name'],
            email=request.form['email'],
            password=generate_password_hash(request.form['password'])))
        db.session.commit()
        flash('Admin account created!', 'success')
        return redirect(url_for('login') + '?tab=admin')
    return render_template('admin_create.html')

# ============================================================
#  ADMIN — DEGREE MANAGEMENT
# ============================================================
@app.route('/admin/degrees')
@login_required
@admin_required
def admin_degrees():
    degrees = Degree.query.order_by(Degree.name).all()
    student_counts = {d.id: Student.query.filter_by(degree=d.name).count()
                      for d in degrees}
    return render_template('admin_degrees.html',
        admin=current_user, degrees=degrees, student_counts=student_counts)

@app.route('/admin/degrees/new', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_degree_new():
    levels = ['UG', 'PG', 'UG Engineering', 'Competitive Exam']
    if request.method == 'POST':
        name = request.form['name'].strip()
        full_name = request.form['full_name'].strip()
        level = request.form.get('level', 'UG')
        if not name or not full_name:
            flash('Name and Full Name are required.', 'error')
            return redirect(url_for('admin_degree_new'))
        if Degree.query.filter_by(name=name).first():
            flash(f'Degree "{name}" already exists.', 'error')
            return redirect(url_for('admin_degree_new'))
        d = Degree(name=name, full_name=full_name, level=level)
        db.session.add(d)
        db.session.commit()
        reload_degree_data()
        flash(f'Degree "{name}" created. Now add roles and skills.', 'success')
        return redirect(url_for('admin_degree_edit', did=d.id))
    return render_template('admin_degree_form.html',
        admin=current_user, degree=None, levels=levels, student_count=0)

@app.route('/admin/degrees/<int:did>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_degree_edit(did):
    d = Degree.query.get_or_404(did)
    levels = ['UG', 'PG', 'UG Engineering', 'Competitive Exam']
    if request.method == 'POST':
        new_name = request.form['name'].strip()
        new_full = request.form['full_name'].strip()
        new_lvl = request.form.get('level', d.level)
        if not new_name or not new_full:
            flash('Name and Full Name are required.', 'error')
            return redirect(url_for('admin_degree_edit', did=did))
        if new_name != d.name:
            if Degree.query.filter_by(name=new_name).first():
                flash(f'Another degree with name "{new_name}" already exists.', 'error')
                return redirect(url_for('admin_degree_edit', did=did))
            # Cascade-rename for students
            Student.query.filter_by(degree=d.name).update({'degree': new_name})
        d.name = new_name
        d.full_name = new_full
        d.level = new_lvl
        db.session.commit()
        reload_degree_data()
        flash(f'Degree "{d.name}" updated.', 'success')
        return redirect(url_for('admin_degree_edit', did=did))
    return render_template('admin_degree_form.html',
        admin=current_user, degree=d, levels=levels,
        student_count=Student.query.filter_by(degree=d.name).count())

@app.route('/admin/degrees/<int:did>/delete', methods=['POST'])
@login_required
@admin_required
def admin_degree_delete(did):
    d = Degree.query.get_or_404(did)
    count = Student.query.filter_by(degree=d.name).count()
    if count > 0:
        flash(f'Cannot delete "{d.name}" — {count} student(s) are registered with this degree.', 'error')
        return redirect(url_for('admin_degrees'))
    name = d.name
    db.session.delete(d)
    db.session.commit()
    reload_degree_data()
    flash(f'Degree "{name}" deleted.', 'success')
    return redirect(url_for('admin_degrees'))

@app.route('/admin/degrees/<int:did>/role/add', methods=['POST'])
@login_required
@admin_required
def admin_role_add(did):
    Degree.query.get_or_404(did)
    name = request.form.get('role', '').strip()
    if name:
        if not DegreeJobRole.query.filter_by(degree_id=did, name=name).first():
            db.session.add(DegreeJobRole(degree_id=did, name=name))
            db.session.commit()
            reload_degree_data()
            flash(f'Role "{name}" added.', 'success')
        else:
            flash(f'Role "{name}" already exists for this degree.', 'info')
    return redirect(url_for('admin_degree_edit', did=did))

@app.route('/admin/degrees/<int:did>/role/<int:rid>/delete', methods=['POST'])
@login_required
@admin_required
def admin_role_delete(did, rid):
    role = DegreeJobRole.query.filter_by(id=rid, degree_id=did).first_or_404()
    name = role.name
    db.session.delete(role)
    db.session.commit()
    reload_degree_data()
    flash(f'Role "{name}" removed.', 'success')
    return redirect(url_for('admin_degree_edit', did=did))

@app.route('/admin/degrees/<int:did>/skill/add', methods=['POST'])
@login_required
@admin_required
def admin_skill_add(did):
    Degree.query.get_or_404(did)
    name = request.form.get('skill', '').strip()
    if name:
        if not DegreeSkill.query.filter_by(degree_id=did, name=name).first():
            db.session.add(DegreeSkill(degree_id=did, name=name))
            db.session.commit()
            reload_degree_data()
            flash(f'Skill "{name}" added.', 'success')
        else:
            flash(f'Skill "{name}" already exists for this degree.', 'info')
    return redirect(url_for('admin_degree_edit', did=did))

@app.route('/admin/degrees/<int:did>/skill/<int:sid>/delete', methods=['POST'])
@login_required
@admin_required
def admin_skill_delete(did, sid):
    skill = DegreeSkill.query.filter_by(id=sid, degree_id=did).first_or_404()
    name = skill.name
    db.session.delete(skill)
    db.session.commit()
    reload_degree_data()
    flash(f'Skill "{name}" removed.', 'success')
    return redirect(url_for('admin_degree_edit', did=did))
    
# ============================================================
#  RUN
# ============================================================
# Initialize DB at module import so gunicorn also sets up tables
with app.app_context():
    db.create_all()
    seed_degrees_if_empty()
    reload_degree_data()
    print('✅ All tables created!')
    print('📌 students | recruiters | admins | predictions | shortlists | degrees')
    print(f'📚 Loaded {len(ALL_DEGREES)} degrees from database')

if __name__ == '__main__':
    print('🌐 Open: http://127.0.0.1:5000')
    app.run(debug=True)